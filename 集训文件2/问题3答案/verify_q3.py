from __future__ import annotations

import csv
import json
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "outputs" / "q3"
DT = 0.2
ACC = 2000.0
DEC = 3000.0
REF_GAP = 1250.0


def read_csv(name: str):
    with (OUT / name).open("r", encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def main() -> None:
    errors: list[str] = []
    links = {}
    ws_link = openpyxl.load_workbook(
        ROOT / "附件2_轨道连接数据.xlsx", data_only=True, read_only=True
    )["Link"]
    for row in ws_link.iter_rows(min_row=2, values_only=True):
        if int(row[1]) == 1:
            links[int(row[0])] = (float(row[4]), float(row[5]))

    tasks = read_csv("问题3_任务结果.csv")
    trace = read_csv("问题3_OHT逐步运行记录.csv")
    metrics = read_csv("问题3_算法评价指标.csv")
    decisions = json.loads((OUT / "问题3_在线决策日志.json").read_text(encoding="utf-8"))
    slots = json.loads((OUT / "问题3_资源时隙日志.json").read_text(encoding="utf-8"))
    reservations = json.loads((OUT / "问题3_资源预约日志.json").read_text(encoding="utf-8"))
    deadlock = json.loads((OUT / "问题3_防死锁日志.json").read_text(encoding="utf-8"))
    plan = json.loads((OUT / "问题3_方案.json").read_text(encoding="utf-8"))
    audit = json.loads((OUT / "问题3_约束审计.json").read_text(encoding="utf-8"))
    if len(tasks) != 600 or len({r["CommandID"] for r in tasks}) != 600:
        errors.append("任务结果不是600个唯一CommandID")

    carriers = defaultdict(list)
    for row in tasks:
        keys = [
            "InstallTime", "AssignedTime", "VehicleFromArrivedTime", "VehicleAcquireStartTime",
            "VehicleAcquireEndTime", "VehicleDepartedTime", "VehicleToArrivedTime",
            "VehicleDepositStartTime", "VehicleDepositEndTime", "TransferCompletedTime",
        ]
        times = [datetime.fromisoformat(row[k]) for k in keys]
        carriers[row["CarrierID"]].append((times[0], row["CommandID"], times))
        if any(a > b for a, b in zip(times, times[1:])):
            errors.append(f"{row['CommandID']} 时间非单调")
        if abs((times[4] - times[3]).total_seconds() - 8.0) > 1e-9:
            errors.append(f"{row['CommandID']} 取货时间错误")
        if abs((times[8] - times[7]).total_seconds() - 8.0) > 1e-9:
            errors.append(f"{row['CommandID']} 放货时间错误")
        if abs((times[9] - times[0]).total_seconds() - float(row["TransferTime"])) > 1e-6:
            errors.append(f"{row['CommandID']} TransferTime错误")
        if abs((times[1] - times[0]).total_seconds() * 1000 - int(row["WaitAssignTime"])) > 1:
            errors.append(f"{row['CommandID']} WaitAssignTime错误")
    for carrier, chain in carriers.items():
        chain.sort(key=lambda x: (x[0], x[1]))
        for previous, current in zip(chain, chain[1:]):
            if previous[2][9] > current[2][1]:
                errors.append(f"{carrier} Carrier前序未完成即分配")

    for decision in decisions:
        if float(decision["max_visible_release_s"]) > float(decision["decision_time_s"]) + 1e-9:
            errors.append("在线决策读取未来任务")
        if decision.get("future_leak"):
            errors.append("决策日志标记future_leak")
        if "region_pressure" not in decision:
            errors.append("决策日志缺少区域压力")

    selected_mode = plan.get("selected_mode")
    if selected_mode not in {"direct", "microbatch", "pressure", "full"}:
        errors.append(f"未知正式模式: {selected_mode}")
    if selected_mode == "full" and (not slots or not reservations):
        errors.append("full模式缺少时隙或资源预约证据")
    by_resource_slots = defaultdict(list)
    for slot in slots:
        if slot.get("actual_entry_s") is not None:
            by_resource_slots[slot["resource"]].append(slot)
    for resource, used in by_resource_slots.items():
        used.sort(key=lambda x: float(x["actual_entry_s"]))
        for left, right in zip(used, used[1:]):
            if left.get("actual_exit_s") is not None and float(right["actual_entry_s"]) < float(left["actual_exit_s"]) - 1e-9:
                errors.append(f"{resource} 实际时隙重叠")
    if deadlock.get("wait_graph_cycle_count", 0) < 0:
        errors.append("等待图统计错误")

    step_counts = Counter(int(r["StepNo"]) for r in trace)
    if sorted(set(step_counts.values())) != [20]:
        errors.append("存在非20车仿真步")
    if step_counts and sorted(step_counts) != list(range(max(step_counts) + 1)):
        errors.append("StepNo不连续")
    previous_speed = {}
    by_step_link = defaultdict(list)
    for row in trace:
        link_id = int(row["CurrentEdgeID"])
        position = float(row["Position"])
        speed = float(row["Speed"])
        length, vmax = links[link_id]
        if not (-1e-6 <= position <= length + 1e-6):
            errors.append(f"{row['VehicleID']}位置越界")
        if not (-1e-6 <= speed <= vmax + 1e-6):
            errors.append(f"{row['VehicleID']}速度越界")
        vehicle_id = row["VehicleID"]
        if vehicle_id in previous_speed:
            acceleration = (speed - previous_speed[vehicle_id]) / DT
            if acceleration < -DEC - 1e-6 or acceleration > ACC + 1e-6:
                errors.append(f"{vehicle_id}加速度越界")
        previous_speed[vehicle_id] = speed
        by_step_link[(int(row["StepNo"]), link_id)].append((position, vehicle_id))
    min_gap = float("inf")
    for (step, link_id), vehicles in by_step_link.items():
        vehicles.sort()
        for back, front in zip(vehicles, vehicles[1:]):
            gap = front[0] - back[0]
            min_gap = min(min_gap, gap)
            if gap < REF_GAP - 1e-6:
                errors.append(f"Step{step} Link{link_id}间距{gap}")

    wb = openpyxl.load_workbook(OUT / "问题3_结果.xlsx", data_only=True, read_only=True)
    if wb["任务仿真结果"].max_row != 823:
        errors.append("Excel任务模板行数错误")
    if wb["OHT逐步运行记录表"].max_row != len(trace) + 1:
        errors.append("Excel轨迹行数不一致")
    if wb["算法评价指标"].cell(4, 2).value != 600:
        errors.append("Excel问题3指标错误")
    if not audit.get("passed") or audit.get("traffic_violations"):
        errors.append("程序内约束审计未通过")

    QUESTION_NO = 3
    evidence_path = ROOT / "正式复核结果" / "独立连续净空复核证据.json"
    if not evidence_path.exists():
        errors.append("缺少独立连续净空复核证据")
    else:
        evidence = json.loads(evidence_path.read_text(encoding="utf-8"))
        matches = [x for x in evidence.get("audits", []) if f"outputs/q{QUESTION_NO}/" in x.get("trace", "")]
        if len(matches) != 1 or not matches[0].get("passed"):
            errors.append("独立连续净空复核未通过")

    result = {
        "passed": not errors,
        "errors": errors[:100],
        "task_rows": len(tasks),
        "trajectory_rows": len(trace),
        "steps": len(step_counts),
        "vehicles_per_step": sorted(set(step_counts.values())),
        "decision_count": len(decisions),
        "slot_count": len(slots),
        "slot_entry_count": sum(x.get("actual_entry_s") is not None for x in slots),
        "reservation_event_count": len(reservations),
        "deadlock_recovery_count": len(deadlock.get("recovery_events", [])),
        "min_same_link_reference_gap_mm": min_gap,
        "metric_row": metrics[0] if metrics else None,
    }
    print(json.dumps(result, ensure_ascii=False, indent=2))
    if errors:
        raise SystemExit(1)


if __name__ == "__main__":
    main()
