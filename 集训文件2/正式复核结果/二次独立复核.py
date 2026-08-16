from __future__ import annotations

import csv
import heapq
import json
import math
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
DT, VEHICLE_LENGTH, MIN_CLEARANCE = 0.2, 950.0, 300.0
MIN_REF_GAP = VEHICLE_LENGTH + MIN_CLEARANCE
ACC_MAX, DEC_MAX, EPS = 2000.0, 3000.0, 1e-3


def csv_rows(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        return list(csv.DictReader(f))


def load_network():
    ws = openpyxl.load_workbook(
        ROOT / "附件2_轨道连接数据.xlsx", data_only=True, read_only=True
    )["Link"]
    links, outgoing = {}, defaultdict(list)
    for r in ws.iter_rows(min_row=2, values_only=True):
        if int(r[1]) != 1:
            continue
        lid, u, v = int(r[0]), int(r[2]), int(r[3])
        length, vmax = float(r[4]), float(r[5])
        links[lid] = {"u": u, "v": v, "length": length, "vmax": vmax}
        outgoing[u].append((v, length, lid))
    return links, outgoing


def shortest(outgoing, start: int, target: int, cutoff=1250.0):
    if start == target:
        return 0.0, []
    pq, best = [(0.0, start, [])], {start: 0.0}
    while pq:
        d, node, path = heapq.heappop(pq)
        if d != best[node] or d > cutoff:
            continue
        if node == target:
            return d, path
        for nxt, length, lid in outgoing.get(node, []):
            nd = d + length
            if nd < best.get(nxt, math.inf) and nd <= cutoff:
                best[nxt] = nd
                heapq.heappush(pq, (nd, nxt, path + [lid]))
    return math.inf, []


def routed_gap(back, front, links, outgoing):
    bl, fl = int(back["CurrentEdgeID"]), int(front["CurrentEdgeID"])
    bp, fp = float(back["Position"]), float(front["Position"])
    if bl == fl:
        return (fp - bp, [bl]) if fp > bp + EPS else (math.inf, [])
    first_text = back["NextEdgeID"].strip()
    if not first_text:
        return math.inf, []
    first = int(first_text)
    if first == fl:
        return links[bl]["length"] - bp + fp, [bl, fl]
    if first not in links or links[bl]["v"] != links[first]["u"]:
        return math.inf, []
    middle, path = shortest(outgoing, links[first]["v"], links[fl]["u"])
    if math.isinf(middle):
        return math.inf, []
    gap = links[bl]["length"] - bp + links[first]["length"] + middle + fp
    return gap, [bl, first] + path + [fl]


def audit_tasks(task_rows):
    errors, carriers = [], defaultdict(list)
    values = defaultdict(list)
    for r in task_rows:
        keys = [
            "InstallTime", "AssignedTime", "VehicleFromArrivedTime",
            "VehicleAcquireStartTime", "VehicleAcquireEndTime",
            "VehicleDepartedTime", "VehicleToArrivedTime",
            "VehicleDepositStartTime", "VehicleDepositEndTime",
            "TransferCompletedTime",
        ]
        t = [datetime.fromisoformat(r[k]) for k in keys]
        if any(a > b for a, b in zip(t, t[1:])):
            errors.append(f"{r['CommandID']}: 时间非单调")
        if abs((t[4] - t[3]).total_seconds() - 8) > 1e-6:
            errors.append(f"{r['CommandID']}: 取货时间错误")
        if abs((t[8] - t[7]).total_seconds() - 8) > 1e-6:
            errors.append(f"{r['CommandID']}: 放货时间错误")
        transfer = (t[9] - t[0]).total_seconds()
        if abs(transfer - float(r["TransferTime"])) > 1e-6:
            errors.append(f"{r['CommandID']}: TransferTime错误")
        if abs((t[1] - t[0]).total_seconds() * 1000 - float(r["WaitAssignTime"])) > 1.1:
            errors.append(f"{r['CommandID']}: WaitAssignTime错误")
        carriers[r["CarrierID"]].append((t[0], r["CommandID"], t[1], t[9]))
        values["AvgTransferTime"].append(transfer)
        values["AvgAssignTime"].append((t[1] - t[0]).total_seconds())
        values["AvgPickupResponseTime"].append((t[2] - t[1]).total_seconds())
        values["AvgTransportTime"].append((t[6] - t[5]).total_seconds())
        values["AvgPausingTime"].append(float(r["PausingTime"]) / 1000)
        de, dl = float(r["ToSourceDistance"]), float(r["ToDestinationDistance"])
        values["AvgEmptyDistance"].append(de)
        values["AvgTotalDistance"].append(de + dl)
        values["completed"].append(t[9])
        values["installed"].append(t[0])
    for carrier, chain in carriers.items():
        chain.sort(key=lambda x: (x[0], x[1]))
        for prev, cur in zip(chain, chain[1:]):
            if prev[3] > cur[2]:
                errors.append(f"{carrier}: 前序完成前已分配后继")
    metrics = {
        k: sum(v) / len(v) for k, v in values.items()
        if k not in {"completed", "installed"}
    }
    metrics["Makespan"] = (
        max(values["completed"]) - min(values["installed"])
    ).total_seconds()
    return errors, metrics


def audit_trace(q: int, path: Path, links, outgoing):
    errors, qnos, step_counts, prev = [], Counter(), Counter(), {}
    min_same, min_routed = (math.inf, None), (math.inf, None)
    max_disp_error = (0.0, None)
    braking_distance_violations = []
    current_step, step_rows = None, []

    def inspect_step(step, items):
        nonlocal min_same, min_routed
        by_link = defaultdict(list)
        for r in items:
            by_link[int(r["CurrentEdgeID"])].append(r)
        for lid, vehicles in by_link.items():
            vehicles.sort(key=lambda x: float(x["Position"]))
            for back, front in zip(vehicles, vehicles[1:]):
                gap = float(front["Position"]) - float(back["Position"])
                if gap < min_same[0]:
                    min_same = (gap, [step, back["VehicleID"], front["VehicleID"], [lid]])
        for back in items:
            for front in items:
                if back["VehicleID"] == front["VehicleID"]:
                    continue
                gap, route = routed_gap(back, front, links, outgoing)
                if 0 < gap < min_routed[0]:
                    min_routed = (gap, [step, back["VehicleID"], front["VehicleID"], route])

    with path.open("r", encoding="utf-8-sig", newline="") as f:
        for r in csv.DictReader(f):
            step, vid = int(r["StepNo"]), r["VehicleID"]
            qnos[int(r["QuestionNo"])] += 1
            step_counts[step] += 1
            if current_step is None:
                current_step = step
            if step != current_step:
                inspect_step(current_step, step_rows)
                current_step, step_rows = step, []
            step_rows.append(r)
            lid, pos, speed = int(r["CurrentEdgeID"]), float(r["Position"]), float(r["Speed"])
            if lid not in links or not (-EPS <= pos <= links[lid]["length"] + EPS):
                errors.append(f"{vid} Step{step}: 位置越界")
            if lid in links and not (-EPS <= speed <= links[lid]["vmax"] + EPS):
                errors.append(f"{vid} Step{step}: 速度越界")
            if vid in prev:
                p = prev[vid]
                acc = (speed - p["speed"]) / DT
                if acc > ACC_MAX + 1e-5 or acc < -DEC_MAX - 1e-5:
                    errors.append(f"{vid} Step{step}: 表观加速度{acc:.3f}越界")
                actual = math.nan
                if p["lid"] == lid:
                    actual = pos - p["pos"]
                elif links[p["lid"]]["v"] == links[lid]["u"]:
                    actual = links[p["lid"]]["length"] - p["pos"] + pos
                trapezoid = 0.5 * (p["speed"] + speed) * DT
                if not math.isnan(actual) and abs(actual - trapezoid) > max_disp_error[0]:
                    max_disp_error = (
                        abs(actual - trapezoid),
                        [step, vid, actual, trapezoid, p["lid"], lid, p["speed"], speed],
                    )
                if not math.isnan(actual) and speed <= EPS and p["speed"] > EPS:
                    min_stop = p["speed"] * p["speed"] / (2 * DEC_MAX)
                    if actual < min_stop - 0.02:
                        braking_distance_violations.append({
                            "step": step,
                            "vehicle": vid,
                            "actual_displacement_mm": actual,
                            "minimum_stopping_distance_mm": min_stop,
                            "shortfall_mm": min_stop - actual,
                            "start_speed_mm_s": p["speed"],
                            "from_link": p["lid"],
                            "to_link": lid,
                        })
            prev[vid] = {"lid": lid, "pos": pos, "speed": speed}
        if step_rows:
            inspect_step(current_step, step_rows)
    if set(qnos) != {q}:
        errors.append(f"QuestionNo错误: {dict(qnos)}")
    if set(step_counts.values()) != {20}:
        errors.append("并非每步20辆车")
    if sorted(step_counts) != list(range(max(step_counts) + 1)):
        errors.append("StepNo不连续")
    return {
        "errors": errors[:500],
        "question_no_counts": dict(qnos),
        "steps": len(step_counts),
        "min_same_link_reference_gap_mm": min_same,
        "min_routed_reference_gap_mm": min_routed,
        "min_routed_clearance_mm": min_routed[0] - VEHICLE_LENGTH,
        "max_trapezoid_displacement_error_mm": max_disp_error,
        "braking_distance_violation_count": len(braking_distance_violations),
        "worst_braking_distance_violations": sorted(
            braking_distance_violations,
            key=lambda x: x["shortfall_mm"],
            reverse=True,
        )[:20],
    }


def p17_audit(links):
    ws = openpyxl.load_workbook(
        ROOT / "附件3_Port位置数据.xlsx", data_only=True, read_only=True
    )["Port"]
    actual = {}
    for r in ws.iter_rows(min_row=2, values_only=True):
        if str(r[0]).upper() == "P17":
            actual = {"node": int(r[1]), "link": int(r[2]), "offset": float(r[3])}
            break
    expected = {"node": 70, "link": 507, "offset": 303.0}
    return {
        "actual": actual,
        "expected": expected,
        "link507": links.get(507),
        "passed": actual == expected and links[507]["u"] == 70
        and 0 <= actual["offset"] <= links[507]["length"],
    }


def main():
    links, outgoing = load_network()
    result = {"p17": p17_audit(links), "questions": {}}
    for q, expected in {1: 32, 2: 190, 3: 600}.items():
        folder = ROOT / "outputs" / f"q{q}"
        tasks = csv_rows(folder / f"问题{q}_任务结果.csv")
        metrics = csv_rows(folder / f"问题{q}_算法评价指标.csv")[0]
        task_errors, recomputed = audit_tasks(tasks)
        result["questions"][str(q)] = {
            "task_count": len(tasks),
            "expected_task_count": expected,
            "task_errors": task_errors,
            "metric_difference_recomputed_minus_csv": {
                k: recomputed[k] - float(metrics[k]) for k in recomputed
            },
            "trace": audit_trace(
                q, folder / f"问题{q}_OHT逐步运行记录.csv", links, outgoing
            ),
        }
    output = Path(__file__).with_name("二次独立复核证据.json")
    output.write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(result, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
