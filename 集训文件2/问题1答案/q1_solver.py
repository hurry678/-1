from __future__ import annotations

import argparse
import copy
import csv
import heapq
import json
import math
import random
import statistics
import time
from collections import defaultdict
from dataclasses import asdict, dataclass, field
from datetime import datetime, timedelta
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Sequence, Tuple

import openpyxl


DT = 0.2
ACC = 2000.0
DEC = 3000.0
VEHICLE_LENGTH = 950.0
CLEAR_GAP = 300.0
REF_GAP = VEHICLE_LENGTH + CLEAR_GAP
PICK_TIME = 8.0
DROP_TIME = 8.0
# 统一采用网络最低限速作为第一版保守巡航上限。这样在任一 Link 切换处均无需
# 瞬时降速；正式审计仍逐 Link 检查附件 2 的实际限速。
CONTROL_VMAX = 710.0
ENTRY_GAP = REF_GAP + CONTROL_VMAX * DT + CONTROL_VMAX * CONTROL_VMAX / (2.0 * DEC) + 100.0
STOP_LINE_BUFFER = 10.0
EPS = 1e-7


@dataclass(frozen=True)
class Node:
    node_id: int
    role: str


@dataclass(frozen=True)
class Link:
    link_id: int
    from_node: int
    to_node: int
    length: float
    vmax: float
    track_class: str
    curve_group: Optional[str]


@dataclass(frozen=True)
class Port:
    port_id: str
    node_id: int
    link_id: int
    offset: float


@dataclass(frozen=True)
class Task:
    command_id: str
    install_time: datetime
    source: str
    destination: str
    priority: int
    carrier_id: str


@dataclass(frozen=True)
class Segment:
    link_id: int
    start: float
    end: float

    @property
    def distance(self) -> float:
        return self.end - self.start


@dataclass
class TaskRecord:
    task: Task
    vehicle_id: str
    assigned_s: float = 0.0
    source_arrived_s: Optional[float] = None
    acquire_start_s: Optional[float] = None
    acquire_end_s: Optional[float] = None
    departed_s: Optional[float] = None
    destination_arrived_s: Optional[float] = None
    deposit_start_s: Optional[float] = None
    deposit_end_s: Optional[float] = None
    completed_s: Optional[float] = None
    to_source_distance: float = 0.0
    to_destination_distance: float = 0.0
    move_to_source_path: str = ""
    move_to_destination_path: str = ""
    paused_count: int = 0
    pausing_ms: int = 0


@dataclass
class Vehicle:
    vehicle_id: str
    link_id: int
    offset: float
    queue: List[str]
    speed: float = 0.0
    acceleration: float = 0.0
    state: str = "IDLE"
    current_task: Optional[str] = None
    carrier_id: Optional[str] = None
    route: List[Segment] = field(default_factory=list)
    route_index: int = 0
    service_remaining: float = 0.0
    target_port: Optional[str] = None
    target_kind: Optional[str] = None
    pause_active: bool = False
    pause_reason: str = ""
    last_reason: str = "正常运行"
    last_action: str = "待命"
    related_vehicle: str = ""
    idle_links_traversed: int = 0


class DataBundle:
    def __init__(self, root: Path):
        self.root = root
        self.nodes: Dict[int, Node] = {}
        self.links: Dict[int, Link] = {}
        self.ports: Dict[str, Port] = {}
        self.vehicles: Dict[str, Tuple[int, float]] = {}
        self.tasks: Dict[str, Task] = {}
        self.curve_members: Dict[str, set[int]] = {}
        self.merge_nodes: set[int] = set()
        self.audit: Dict[str, object] = {}

    @staticmethod
    def _rows(path: Path, sheet: str) -> List[Tuple[object, ...]]:
        ws = openpyxl.load_workbook(path, data_only=True, read_only=True)[sheet]
        return list(ws.iter_rows(values_only=True))

    def load(self) -> "DataBundle":
        for row in self._rows(self.root / "附件1_轨道节点数据.xlsx", "Node")[1:]:
            node_id, _, _, role = row
            self.nodes[int(node_id)] = Node(int(node_id), str(role))

        for row in self._rows(self.root / "附件2_轨道连接数据.xlsx", "Link")[1:]:
            link_id, use_flag, f, t, length, vmax, kind, group = row
            if int(use_flag) != 1:
                continue
            self.links[int(link_id)] = Link(
                int(link_id), int(f), int(t), float(length), float(vmax),
                str(kind), None if group is None else str(group),
            )

        for row in self._rows(self.root / "附件3_Port位置数据.xlsx", "Port")[1:]:
            port_id, node_id, link_id, offset = row
            self.ports[str(port_id)] = Port(str(port_id), int(node_id), int(link_id), float(offset))

        # 用户提供的全过程计划要求在构图和缓存前执行该唯一白名单修订。
        self.ports["P17"] = Port("P17", 70, 507, 303.0)

        for row in self._rows(self.root / "附件4_OHT初始位置数据.xlsx", "Vehicle")[1:]:
            vehicle_id, _, link_id, offset = row
            self.vehicles[str(vehicle_id)] = (int(link_id), float(offset))

        for row in self._rows(self.root / "附件5_静态任务数据.xlsx", "Task")[1:]:
            command_id, install_time, source, destination, priority, carrier_id = row
            task = Task(str(command_id), install_time, str(source), str(destination), int(priority), str(carrier_id))
            self.tasks[task.command_id] = task

        for row in self._rows(self.root / "附件8_运行参数与控制对象.xlsx", "ControlledObjects")[1:]:
            kind, control_id, capacity, members = row
            if int(capacity) != 1:
                raise ValueError(f"控制对象 {control_id} 容量不是 1")
            if kind == "CURVE":
                self.curve_members[str(control_id)] = {int(x) for x in str(members).split(";")}
            elif kind == "MERGE":
                self.merge_nodes.add(int(str(control_id).lstrip("N")))

        self._audit()
        return self

    def _audit(self) -> None:
        errors: List[str] = []
        for link in self.links.values():
            if link.from_node not in self.nodes or link.to_node not in self.nodes:
                errors.append(f"Link{link.link_id} 端点外键无效")
            if link.length <= 0 or link.vmax <= 0:
                errors.append(f"Link{link.link_id} 长度或限速无效")
        for port in self.ports.values():
            link = self.links.get(port.link_id)
            if link is None:
                errors.append(f"{port.port_id} Link 外键无效")
            elif port.node_id != link.from_node or not (-EPS <= port.offset <= link.length + EPS):
                errors.append(f"{port.port_id} 位置非法")
        for vehicle_id, (link_id, offset) in self.vehicles.items():
            link = self.links.get(link_id)
            if link is None or not (-EPS <= offset <= link.length + EPS):
                errors.append(f"{vehicle_id} 初始位置非法")
        for task in self.tasks.values():
            if task.source not in self.ports or task.destination not in self.ports:
                errors.append(f"{task.command_id} Port 外键无效")
        for group, members in self.curve_members.items():
            actual = {x.link_id for x in self.links.values() if x.curve_group == group}
            if actual != members:
                errors.append(f"{group} 成员不一致: link={sorted(actual)}, control={sorted(members)}")
        expected = {
            "nodes": 131, "links": 152, "ports": 36, "vehicles": 20,
            "tasks": 32, "curve_groups": 33, "merges": 21,
        }
        actual_counts = {
            "nodes": len(self.nodes), "links": len(self.links), "ports": len(self.ports),
            "vehicles": len(self.vehicles), "tasks": len(self.tasks),
            "curve_groups": len(self.curve_members), "merges": len(self.merge_nodes),
        }
        for key, value in expected.items():
            if actual_counts[key] != value:
                errors.append(f"{key} 数量 {actual_counts[key]} != {value}")
        p17 = self.ports["P17"]
        if (p17.node_id, p17.link_id, p17.offset) != (70, 507, 303.0):
            errors.append("P17 强制修订未生效")
        self.audit = {
            "counts": actual_counts,
            "p17": asdict(p17),
            "clear_gap_mm": CLEAR_GAP,
            "reference_gap_mm": REF_GAP,
            "errors": errors,
            "passed": not errors,
        }
        if errors:
            raise ValueError("数据审计失败:\n" + "\n".join(errors))


class GraphEngine:
    def __init__(self, data: DataBundle):
        self.data = data
        self.adj: Dict[int, List[Link]] = defaultdict(list)
        for link in data.links.values():
            if data.nodes[link.to_node].role == "TERMINAL":
                continue
            self.adj[link.from_node].append(link)
        for links in self.adj.values():
            links.sort(key=lambda x: x.link_id)
        self._path_cache: Dict[Tuple[int, int], List[int]] = {}
        self._position_cache: Dict[Tuple[int, int, int, int], Tuple[List[Segment], float, float, str]] = {}

    @staticmethod
    def _segment_free_time(distance: float, vmax: float) -> float:
        if distance <= EPS:
            return 0.0
        # 从静止到限速的下界近似；调度代理使用，正式时间由运动仿真产生。
        d_acc = vmax * vmax / (2.0 * ACC)
        d_dec = vmax * vmax / (2.0 * DEC)
        if distance >= d_acc + d_dec:
            return vmax / ACC + vmax / DEC + (distance - d_acc - d_dec) / vmax
        peak = math.sqrt(2.0 * distance / (1.0 / ACC + 1.0 / DEC))
        return peak / ACC + peak / DEC

    def shortest_links(self, start_node: int, target_node: int) -> List[int]:
        key = (start_node, target_node)
        if key in self._path_cache:
            return list(self._path_cache[key])
        if start_node == target_node:
            self._path_cache[key] = []
            return []
        dist: Dict[int, float] = {start_node: 0.0}
        prev: Dict[int, Tuple[int, int]] = {}
        heap: List[Tuple[float, int]] = [(0.0, start_node)]
        while heap:
            d, node = heapq.heappop(heap)
            if d > dist.get(node, math.inf) + EPS:
                continue
            if node == target_node:
                break
            for link in self.adj.get(node, []):
                nd = d + link.length / link.vmax
                if nd + EPS < dist.get(link.to_node, math.inf):
                    dist[link.to_node] = nd
                    prev[link.to_node] = (node, link.link_id)
                    heapq.heappush(heap, (nd, link.to_node))
        if target_node not in dist:
            raise ValueError(f"不存在从 Node{start_node} 到 Node{target_node} 的正常任务路径")
        result: List[int] = []
        node = target_node
        while node != start_node:
            node0, link_id = prev[node]
            result.append(link_id)
            node = node0
        result.reverse()
        self._path_cache[key] = result
        return list(result)

    def position_path(self, start_link: int, start_offset: float, target: Port) -> Tuple[List[Segment], float, float, str]:
        key = (start_link, int(round(start_offset * 10)), target.link_id, int(round(target.offset * 10)))
        if key in self._position_cache:
            segs, dist, tt, nodes = self._position_cache[key]
            return list(segs), dist, tt, nodes
        first = self.data.links[start_link]
        last = self.data.links[target.link_id]
        segments: List[Segment] = []
        if start_link == target.link_id and target.offset >= start_offset - EPS:
            segments.append(Segment(start_link, start_offset, target.offset))
        else:
            segments.append(Segment(start_link, start_offset, first.length))
            middle = self.shortest_links(first.to_node, last.from_node)
            for link_id in middle:
                link = self.data.links[link_id]
                segments.append(Segment(link_id, 0.0, link.length))
            segments.append(Segment(last.link_id, 0.0, target.offset))
        segments = [s for s in segments if s.distance > EPS or len(segments) == 1]
        distance = sum(s.distance for s in segments)
        free_time = sum(self._segment_free_time(s.distance, self.data.links[s.link_id].vmax) for s in segments)
        node_seq: List[int] = [self.data.links[segments[0].link_id].from_node]
        for segment in segments:
            link = self.data.links[segment.link_id]
            if abs(segment.end - link.length) <= EPS:
                if node_seq[-1] != link.to_node:
                    node_seq.append(link.to_node)
            elif node_seq[-1] != link.from_node:
                node_seq.append(link.from_node)
        nodes = ";".join(str(x) for x in node_seq)
        self._position_cache[key] = (list(segments), distance, free_time, nodes)
        return list(segments), distance, free_time, nodes

    def idle_next_link(self, current_link: int) -> int:
        node = self.data.links[current_link].to_node
        options = self.adj.get(node, [])
        if not options:
            raise RuntimeError(f"Node{node} 无合法非终点后继 Link")
        # 空闲车优先走直轨、较高限速 Link，减少其对任务车的阻塞。
        return min(options, key=lambda x: (x.track_class == "CURVE", -x.vmax, x.link_id)).link_id


class StaticScheduler:
    def __init__(self, data: DataBundle, graph: GraphEngine, seed: int = 20260815):
        self.data = data
        self.graph = graph
        self.rng = random.Random(seed)
        self.vehicle_ids = sorted(data.vehicles)
        self.task_ids = list(data.tasks)
        self._leg_cache: Dict[Tuple[str, str, str], Tuple[float, float]] = {}

    def _travel(self, origin: Tuple[int, float] | str, port_id: str) -> Tuple[float, float]:
        if isinstance(origin, tuple):
            key = (f"L{origin[0]}@{origin[1]:.1f}", port_id, "v")
            start_link, offset = origin
        else:
            key = (origin, port_id, "p")
            p = self.data.ports[origin]
            start_link, offset = p.link_id, p.offset
        if key not in self._leg_cache:
            _, distance, tt, _ = self.graph.position_path(start_link, offset, self.data.ports[port_id])
            self._leg_cache[key] = (distance, tt)
        return self._leg_cache[key]

    def evaluate(self, chains: Dict[str, List[str]]) -> Tuple[float, Dict[str, float]]:
        total_completion = 0.0
        priority_delay = 0.0
        empty_distance = 0.0
        makespan = 0.0
        for vehicle_id in self.vehicle_ids:
            origin: Tuple[int, float] | str = self.data.vehicles[vehicle_id]
            elapsed = 0.0
            for task_id in chains[vehicle_id]:
                task = self.data.tasks[task_id]
                de, te = self._travel(origin, task.source)
                dl, tl = self._travel(task.source, task.destination)
                elapsed += te + PICK_TIME + tl + DROP_TIME
                total_completion += elapsed
                priority_delay += max(0, task.priority - 50) * elapsed
                empty_distance += de
                origin = task.destination
            makespan = max(makespan, elapsed)
        score = total_completion + 0.002 * priority_delay + 0.03 * makespan + 1e-5 * empty_distance
        return score, {
            "surrogate_sum_completion": total_completion,
            "surrogate_makespan": makespan,
            "surrogate_empty_distance": empty_distance,
            "surrogate_score": score,
        }

    def baseline(self) -> Dict[str, List[str]]:
        chains = {v: [] for v in self.vehicle_ids}
        tasks = sorted(self.data.tasks.values(), key=lambda t: (-t.priority, t.command_id))
        for task in tasks:
            best_vehicle = None
            best_score = math.inf
            for vehicle_id in self.vehicle_ids:
                trial = {v: list(q) for v, q in chains.items()}
                trial[vehicle_id].append(task.command_id)
                score, _ = self.evaluate(trial)
                if score < best_score - EPS:
                    best_score, best_vehicle = score, vehicle_id
            assert best_vehicle is not None
            chains[best_vehicle].append(task.command_id)
        return chains

    def insertion_initial(self) -> Dict[str, List[str]]:
        chains = {v: [] for v in self.vehicle_ids}
        tasks = sorted(self.data.tasks.values(), key=lambda t: (-t.priority, t.command_id))
        for task in tasks:
            best = None
            for vehicle_id in self.vehicle_ids:
                for pos in range(len(chains[vehicle_id]) + 1):
                    trial = {v: list(q) for v, q in chains.items()}
                    trial[vehicle_id].insert(pos, task.command_id)
                    score, _ = self.evaluate(trial)
                    candidate = (score, vehicle_id, pos)
                    if best is None or candidate < best:
                        best = candidate
            assert best is not None
            chains[best[1]].insert(best[2], task.command_id)
        return chains

    def _remove(self, chains: Dict[str, List[str]], count: int, mode: str) -> Tuple[Dict[str, List[str]], List[str]]:
        result = {v: list(q) for v, q in chains.items()}
        all_ids = [task for q in result.values() for task in q]
        if mode == "random":
            removed = self.rng.sample(all_ids, count)
        elif mode == "priority_cluster":
            pivot = self.data.tasks[self.rng.choice(all_ids)]
            removed = sorted(
                all_ids,
                key=lambda x: (
                    self.data.tasks[x].source != pivot.source,
                    self.data.tasks[x].destination != pivot.destination,
                    abs(self.data.tasks[x].priority - pivot.priority),
                ),
            )[:count]
        else:
            base, _ = self.evaluate(result)
            contributions = []
            for task_id in all_ids:
                trial = {v: [x for x in q if x != task_id] for v, q in result.items()}
                score, _ = self.evaluate(trial)
                contributions.append((base - score, task_id))
            removed = [x[1] for x in sorted(contributions, reverse=True)[:count]]
        removed_set = set(removed)
        for vehicle_id in result:
            result[vehicle_id] = [x for x in result[vehicle_id] if x not in removed_set]
        return result, removed

    def _repair(self, chains: Dict[str, List[str]], removed: List[str]) -> Dict[str, List[str]]:
        result = {v: list(q) for v, q in chains.items()}
        pending = list(removed)
        while pending:
            choices = []
            for task_id in pending:
                insertions = []
                for vehicle_id in self.vehicle_ids:
                    for pos in range(len(result[vehicle_id]) + 1):
                        trial = {v: list(q) for v, q in result.items()}
                        trial[vehicle_id].insert(pos, task_id)
                        score, _ = self.evaluate(trial)
                        insertions.append((score, vehicle_id, pos))
                insertions.sort()
                regret = insertions[min(2, len(insertions) - 1)][0] - insertions[0][0]
                choices.append((-regret, insertions[0], task_id))
            _, best_insert, task_id = min(choices)
            result[best_insert[1]].insert(best_insert[2], task_id)
            pending.remove(task_id)
        return result

    def alns(self, iterations: int = 1200) -> Tuple[Dict[str, List[str]], Dict[str, object]]:
        current = self.insertion_initial()
        current_score, _ = self.evaluate(current)
        best = copy.deepcopy(current)
        best_score = current_score
        temperature = max(1.0, current_score * 0.02)
        accepted = 0
        improved = 0
        modes = ["random", "worst", "priority_cluster"]
        for iteration in range(iterations):
            count = self.rng.randint(2, 6)
            partial, removed = self._remove(current, count, modes[iteration % len(modes)])
            candidate = self._repair(partial, removed)
            score, _ = self.evaluate(candidate)
            delta = score - current_score
            if delta <= 0 or self.rng.random() < math.exp(-delta / max(temperature, 1e-9)):
                current, current_score = candidate, score
                accepted += 1
            if score < best_score - EPS:
                best, best_score = copy.deepcopy(candidate), score
                improved += 1
            temperature *= 0.995
        details = {
            "algorithm": "surrogate_ALNS",
            "iterations": iterations,
            "seed": 20260815,
            "accepted": accepted,
            "improvements": improved,
            "best_surrogate": self.evaluate(best)[1],
        }
        return best, details


class TrafficSimulator:
    MOVING_STATES = {"TO_PICKUP", "TO_DROPOFF", "IDLE"}

    def __init__(self, data: DataBundle, graph: GraphEngine, chains: Dict[str, List[str]], label: str):
        self.data = data
        self.graph = graph
        self.chains = {v: list(q) for v, q in chains.items()}
        self.label = label
        self.time_s = 0.0
        self.step_no = 0
        self.vehicles: Dict[str, Vehicle] = {}
        self.records: Dict[str, TaskRecord] = {}
        self.port_owner: Dict[str, str] = {}
        self.curve_owner: Dict[str, str] = {}
        self.curve_release: Dict[str, Tuple[str, int, float]] = {}
        self.merge_owner: Dict[int, str] = {}
        self.merge_release: Dict[int, Tuple[str, int, float]] = {}
        self.trajectory: List[List[object]] = []
        self.violations: List[Dict[str, object]] = []
        self.max_observed_speed = 0.0
        self.max_observed_acc = 0.0
        self.min_observed_acc = 0.0
        self.min_same_link_gap = math.inf
        self.completed_count = 0
        self._initialize()

    def _initialize(self) -> None:
        assigned = set()
        for vehicle_id in sorted(self.data.vehicles):
            link_id, offset = self.data.vehicles[vehicle_id]
            queue = list(self.chains.get(vehicle_id, []))
            vehicle = Vehicle(vehicle_id, link_id, offset, queue)
            self.vehicles[vehicle_id] = vehicle
            for task_id in queue:
                if task_id in assigned:
                    raise ValueError(f"任务重复分配: {task_id}")
                assigned.add(task_id)
                self.records[task_id] = TaskRecord(self.data.tasks[task_id], vehicle_id)
        if assigned != set(self.data.tasks):
            raise ValueError("调度方案没有完整覆盖 32 项任务")
        for vehicle in self.vehicles.values():
            self._activate_next_task(vehicle, 0.0)
        self._handle_arrivals(0.0)
        self._audit_state(0.0)
        self._log_trajectory()

    def _route_to_port(self, vehicle: Vehicle, port_id: str, kind: str) -> None:
        port = self.data.ports[port_id]
        route, distance, _, nodes = self.graph.position_path(vehicle.link_id, vehicle.offset, port)
        vehicle.route = route
        vehicle.route_index = 0
        vehicle.target_port = port_id
        vehicle.target_kind = kind
        record = self.records[vehicle.current_task]
        if kind == "pickup":
            record.to_source_distance = distance
            record.move_to_source_path = nodes
            vehicle.state = "TO_PICKUP"
        else:
            record.to_destination_distance = distance
            record.move_to_destination_path = nodes
            vehicle.state = "TO_DROPOFF"
        vehicle.last_action = "运行"
        vehicle.last_reason = "正常运行"

    def _activate_next_task(self, vehicle: Vehicle, at_s: float) -> None:
        if vehicle.queue:
            task_id = vehicle.queue.pop(0)
            vehicle.current_task = task_id
            vehicle.carrier_id = None
            task = self.data.tasks[task_id]
            self._route_to_port(vehicle, task.source, "pickup")
        else:
            vehicle.current_task = None
            vehicle.carrier_id = None
            vehicle.state = "IDLE"
            vehicle.target_port = None
            vehicle.target_kind = None
            # 任务路径的末段止于 Link 中部 Port。转入空闲巡航时必须重新建立
            # “当前位置→当前 Link 终点”的剩余段，不能沿用止于 Port 的旧末段。
            vehicle.route = []
            vehicle.route_index = 0
            self._ensure_idle_route(vehicle)

    def _ensure_idle_route(self, vehicle: Vehicle, minimum_future: int = 6) -> None:
        if vehicle.route and vehicle.route_index < len(vehicle.route):
            route = vehicle.route[:vehicle.route_index + 1]
        else:
            link = self.data.links[vehicle.link_id]
            route = [Segment(vehicle.link_id, vehicle.offset, link.length)]
            vehicle.route_index = 0
        current_link = route[-1].link_id
        while len(route) - vehicle.route_index < minimum_future:
            next_link = self.graph.idle_next_link(current_link)
            link = self.data.links[next_link]
            route.append(Segment(next_link, 0.0, link.length))
            current_link = next_link
        vehicle.route = route

    def _current_segment(self, vehicle: Vehicle) -> Segment:
        if vehicle.state == "IDLE":
            self._ensure_idle_route(vehicle)
        if not vehicle.route or vehicle.route_index >= len(vehicle.route):
            raise RuntimeError(f"{vehicle.vehicle_id} 缺少运行路径")
        return vehicle.route[vehicle.route_index]

    def _next_segment(self, vehicle: Vehicle) -> Optional[Segment]:
        if vehicle.state == "IDLE":
            self._ensure_idle_route(vehicle)
        i = vehicle.route_index + 1
        return vehicle.route[i] if i < len(vehicle.route) else None

    def _remaining_distance(self, vehicle: Vehicle) -> float:
        segment = self._current_segment(vehicle)
        value = max(0.0, segment.end - vehicle.offset)
        for seg in vehicle.route[vehicle.route_index + 1:]:
            value += seg.distance
        return value

    def _task_priority(self, vehicle: Vehicle) -> int:
        return self.data.tasks[vehicle.current_task].priority if vehicle.current_task else -1

    def _release_resources(self) -> None:
        for group, (vehicle_id, link_id, offset) in list(self.curve_release.items()):
            vehicle = self.vehicles[vehicle_id]
            if vehicle.link_id != link_id or vehicle.offset + EPS >= offset:
                self.curve_owner.pop(group, None)
                self.curve_release.pop(group, None)
        for node, (vehicle_id, link_id, offset) in list(self.merge_release.items()):
            vehicle = self.vehicles[vehicle_id]
            if vehicle.link_id != link_id or vehicle.offset + EPS >= offset:
                self.merge_owner.pop(node, None)
                self.merge_release.pop(node, None)

    def _request_distance(self, vehicle: Vehicle) -> float:
        segment = self._current_segment(vehicle)
        return max(0.0, segment.end - vehicle.offset)

    def _upcoming_controls(self, vehicle: Vehicle, horizon: float = 6000.0) -> List[Tuple[str, object, float]]:
        """返回路径前方控制边界及沿路径距离，允许在短 Link 之前提前预约。"""
        result: List[Tuple[str, object, float]] = []
        cumulative = 0.0
        upper = min(len(vehicle.route) - 1, vehicle.route_index + 12)
        for idx in range(vehicle.route_index, upper):
            segment = vehicle.route[idx]
            start = vehicle.offset if idx == vehicle.route_index else segment.start
            cumulative += max(0.0, segment.end - start)
            current = self.data.links[segment.link_id]
            next_link = self.data.links[vehicle.route[idx + 1].link_id]
            if current.to_node in self.data.merge_nodes:
                result.append(("MERGE", current.to_node, cumulative))
            if next_link.curve_group and next_link.curve_group != current.curve_group:
                result.append(("CURVE", next_link.curve_group, cumulative))
            if cumulative > horizon:
                break
        return result

    def _arbitrate_resources(self) -> None:
        self._release_resources()
        # 预约尚未生效且车前已有实体时撤销预约，避免队尾车辆持有前方资源、
        # 队首又等待队尾释放资源所形成的循环等待。
        for group, vehicle_id in list(self.curve_owner.items()):
            if group in self.curve_release:
                continue
            vehicle = self.vehicles[vehicle_id]
            controls = self._upcoming_controls(vehicle)
            front = self._front_ahead(vehicle)
            match = next((c for c in controls if c[0] == "CURVE" and str(c[1]) == group), None)
            if match is not None:
                revoke_margin = STOP_LINE_BUFFER + vehicle.speed * DT + vehicle.speed * vehicle.speed / (2.0 * DEC) + 100.0
                if front is not None and front[1] < match[2] + REF_GAP and match[2] > revoke_margin:
                    self.curve_owner.pop(group, None)
        for node, vehicle_id in list(self.merge_owner.items()):
            if node in self.merge_release:
                continue
            vehicle = self.vehicles[vehicle_id]
            controls = self._upcoming_controls(vehicle)
            front = self._front_ahead(vehicle)
            match = next((c for c in controls if c[0] == "MERGE" and int(c[1]) == node), None)
            if match is not None:
                revoke_margin = REF_GAP + vehicle.speed * DT + vehicle.speed * vehicle.speed / (2.0 * DEC) + 100.0
                if front is not None and front[1] < match[2] + REF_GAP and match[2] > revoke_margin:
                    self.merge_owner.pop(node, None)

        curve_requests: Dict[str, List[Tuple[Vehicle, float]]] = defaultdict(list)
        merge_requests: Dict[int, List[Tuple[Vehicle, float]]] = defaultdict(list)
        for vehicle in self.vehicles.values():
            if vehicle.state not in self.MOVING_STATES:
                continue
            lookahead = max(4.0 * REF_GAP, vehicle.speed * vehicle.speed / (2.0 * DEC) + vehicle.speed * 1.0 + REF_GAP)
            for kind, resource, distance in self._upcoming_controls(vehicle, lookahead):
                if distance <= lookahead + EPS:
                    front = self._front_ahead(vehicle)
                    if front is not None and front[1] < distance + REF_GAP:
                        continue
                    if kind == "CURVE":
                        curve_requests[str(resource)].append((vehicle, distance))
                    else:
                        merge_requests[int(resource)].append((vehicle, distance))

        def rank(item: Tuple[Vehicle, float]) -> Tuple[float, int, int, str]:
            v, distance = item
            loaded_rank = 0 if v.state == "TO_DROPOFF" else (1 if v.state == "TO_PICKUP" else 2)
            # 单向轨道禁止超车，因此物理上最靠近停止线的车辆必须先获权；
            # 载货和任务优先级只用于相同 ETA 的仲裁，否则会把权限授予队尾而锁死整列车。
            return distance, loaded_rank, -self._task_priority(v), v.vehicle_id

        for group, requests in curve_requests.items():
            if group not in self.curve_owner:
                winner, _ = min(requests, key=rank)
                self.curve_owner[group] = winner.vehicle_id
        for node, requests in merge_requests.items():
            if node not in self.merge_owner:
                winner, _ = min(requests, key=rank)
                self.merge_owner[node] = winner.vehicle_id

    def _front_ahead(self, vehicle: Vehicle) -> Optional[Tuple[Vehicle, float]]:
        segment = self._current_segment(vehicle)
        best: Optional[Tuple[Vehicle, float]] = None
        cumulative = 0.0
        for idx in range(vehicle.route_index, min(len(vehicle.route), vehicle.route_index + 8)):
            seg = vehicle.route[idx]
            seg_start = vehicle.offset if idx == vehicle.route_index else seg.start
            for other in self.vehicles.values():
                if other.vehicle_id == vehicle.vehicle_id or other.link_id != seg.link_id:
                    continue
                if idx == vehicle.route_index and other.offset <= vehicle.offset + EPS:
                    continue
                if other.offset < seg_start - EPS:
                    continue
                distance = cumulative + other.offset - seg_start
                if distance <= EPS:
                    continue
                if best is None or distance < best[1]:
                    best = (other, distance)
            cumulative += seg.end - seg_start
            if cumulative > 6.0 * REF_GAP:
                break
        return best

    def _boundary_block(self, vehicle: Vehicle) -> Tuple[bool, str, str]:
        current = self.data.links[vehicle.link_id]
        next_seg = self._next_segment(vehicle)
        if next_seg is None:
            return False, "", ""
        next_link = self.data.links[next_seg.link_id]
        if next_link.curve_group and next_link.curve_group != current.curve_group:
            if self.curve_owner.get(next_link.curve_group) != vehicle.vehicle_id:
                owner = self.curve_owner.get(next_link.curve_group, "")
                return True, "弯轨等待", owner
        if current.to_node in self.data.merge_nodes:
            if self.merge_owner.get(current.to_node) != vehicle.vehicle_id:
                owner = self.merge_owner.get(current.to_node, "")
                return True, "汇流避让", owner
        # 下游 Link 上的同向实体由路径连续的 _front_ahead 处理；在这里再按
        # next_link.offset 单独封锁会忽略当前 Link 的剩余距离，并在短 Link 上
        # 造成“突然出现零制动距离”。不同来向的合流由 MERGE 权限处理。
        return False, "", ""

    def _movement_control(self, vehicle: Vehicle) -> Tuple[float, float, str, str]:
        link = self.data.links[vehicle.link_id]
        desired = min(link.vmax, CONTROL_VMAX)
        hard_stop = math.inf
        reason = "正常运行" if vehicle.current_task else "空闲巡航"
        related = ""

        def safe_next_speed(stop_distance: float) -> float:
            """使下一离散步结束后仍保有以 DEC 停车的距离。"""
            d = max(0.0, stop_distance)
            b = DEC * DT
            discriminant = b * b - 4.0 * (DEC * DT * vehicle.speed - 2.0 * DEC * d)
            if discriminant <= 0.0:
                return 0.0
            return max(0.0, (-b + math.sqrt(discriminant)) / 2.0)

        if vehicle.current_task:
            remaining = self._remaining_distance(vehicle)
            hard_stop = min(hard_stop, remaining)
            desired = min(desired, safe_next_speed(remaining))
            if remaining <= max(1.0, vehicle.speed * DT):
                reason = "到达作业点"

        front = self._front_ahead(vehicle)
        if front is not None:
            other, reference_distance = front
            available = reference_distance - REF_GAP
            # 将前车视为可立即停车，得到比同减速度假设更保守的跟驰控制。
            safe_v = safe_next_speed(max(0.0, available))
            if safe_v < desired - EPS:
                desired = safe_v
                hard_stop = min(hard_stop, max(0.0, available))
                reason = "Port等待" if other.state in {"PICKING", "DROPPING"} else "安全跟驰"
                related = other.vehicle_id

        # 对短 Link 后的控制对象提前制动，避免车辆进入短 Link 后才发现已没有
        # 足够的上游安全间距。这里只处理路径上的首个控制边界。
        for kind, resource, distance_to_boundary in self._upcoming_controls(vehicle):
            if kind == "MERGE":
                owner = self.merge_owner.get(int(resource))
                block_reason = "汇流避让"
                buffer = REF_GAP
            else:
                owner = self.curve_owner.get(str(resource))
                block_reason = "弯轨等待"
                buffer = STOP_LINE_BUFFER
            if owner != vehicle.vehicle_id:
                distance = max(0.0, distance_to_boundary - buffer)
                desired = min(desired, safe_next_speed(distance))
                hard_stop = min(hard_stop, distance)
                reason, related = block_reason, owner or ""

        blocked, block_reason, block_related = self._boundary_block(vehicle)
        if blocked:
            # 未获权车辆停在节点上游，而不是以非零速度贴到 Link 端点；留出小量
            # 数值缓冲也避免浮点舍入导致下一步才发现零制动距离。汇流与下游入口
            # 等待需在上游保留完整参考点间距，防止另一支路车辆切入后间距瞬时不足。
            buffer = REF_GAP if block_reason in {"汇流避让", "安全跟驰"} else STOP_LINE_BUFFER
            distance = max(0.0, self._current_segment(vehicle).end - vehicle.offset - buffer)
            desired = min(desired, safe_next_speed(distance))
            hard_stop = min(hard_stop, distance)
            if desired < link.vmax - EPS or distance < 2.0 * REF_GAP:
                reason, related = block_reason, block_related

        desired = max(0.0, desired)
        if math.isfinite(hard_stop):
            desired = min(desired, max(0.0, 2.0 * hard_stop / DT - vehicle.speed))
        acceleration = max(-DEC, min(ACC, (desired - vehicle.speed) / DT))
        new_speed = max(0.0, min(link.vmax, vehicle.speed + acceleration * DT))
        movement = 0.5 * (vehicle.speed + new_speed) * DT
        if math.isfinite(hard_stop) and movement > hard_stop + 1e-6:
            feasible_new_speed = max(0.0, 2.0 * hard_stop / DT - vehicle.speed)
            required_acc = (feasible_new_speed - vehicle.speed) / DT
            if required_acc < -DEC - 1e-5:
                self._violate("braking", vehicle.vehicle_id, f"停止距离不足 {hard_stop:.3f} mm")
            new_speed = feasible_new_speed
            acceleration = required_acc
            movement = hard_stop
        return movement, new_speed, reason, related

    def _cross_boundary(self, vehicle: Vehicle, old_link_id: int, next_link_id: int) -> None:
        old_link = self.data.links[old_link_id]
        next_link = self.data.links[next_link_id]
        if old_link.curve_group and old_link.curve_group != next_link.curve_group:
            group = old_link.curve_group
            if self.curve_owner.get(group) != vehicle.vehicle_id:
                self._violate("curve_owner", vehicle.vehicle_id, f"无权离开/占用 {group}")
            self.curve_release[group] = (vehicle.vehicle_id, next_link_id, min(next_link.length, VEHICLE_LENGTH))
        if next_link.curve_group and next_link.curve_group != old_link.curve_group:
            if self.curve_owner.get(next_link.curve_group) != vehicle.vehicle_id:
                self._violate("curve_entry", vehicle.vehicle_id, f"无权进入 {next_link.curve_group}")
        if old_link.to_node in self.data.merge_nodes:
            node = old_link.to_node
            if self.merge_owner.get(node) != vehicle.vehicle_id:
                self._violate("merge_entry", vehicle.vehicle_id, f"无权通过 Node{node}")
            self.merge_release[node] = (vehicle.vehicle_id, next_link_id, min(next_link.length, REF_GAP))

    def _advance_vehicle(self, vehicle: Vehicle, movement: float, new_speed: float, boundary_allowed: bool) -> None:
        left = movement
        while left > EPS:
            segment = self._current_segment(vehicle)
            available = max(0.0, segment.end - vehicle.offset)
            if left < available - EPS:
                vehicle.offset += left
                left = 0.0
            else:
                vehicle.offset = segment.end
                left -= available
                next_seg = self._next_segment(vehicle)
                if next_seg is None:
                    left = 0.0
                    new_speed = 0.0
                    break
                if not boundary_allowed:
                    if left > 1e-5:
                        self._violate("boundary_cross", vehicle.vehicle_id, "受阻仍试图跨越 Link 边界")
                    left = 0.0
                    new_speed = 0.0
                    break
                old_link_id = vehicle.link_id
                vehicle.route_index += 1
                vehicle.link_id = next_seg.link_id
                vehicle.offset = next_seg.start
                self._cross_boundary(vehicle, old_link_id, next_seg.link_id)
                if vehicle.state == "IDLE":
                    vehicle.idle_links_traversed += 1
                    self._ensure_idle_route(vehicle)
        vehicle.speed = max(0.0, new_speed)

    def _process_services(self, end_s: float) -> None:
        for vehicle in self.vehicles.values():
            if vehicle.state not in {"PICKING", "DROPPING"}:
                continue
            vehicle.service_remaining = max(0.0, vehicle.service_remaining - DT)
            vehicle.speed = 0.0
            vehicle.acceleration = 0.0
            vehicle.last_action = "取货" if vehicle.state == "PICKING" else "放货"
            vehicle.last_reason = "Port作业"
            if vehicle.service_remaining > EPS:
                continue
            task_id = vehicle.current_task
            assert task_id is not None and vehicle.target_port is not None
            record = self.records[task_id]
            port_id = vehicle.target_port
            self.port_owner.pop(port_id, None)
            if vehicle.state == "PICKING":
                record.acquire_end_s = end_s
                record.departed_s = end_s
                vehicle.carrier_id = record.task.carrier_id
                self._route_to_port(vehicle, record.task.destination, "dropoff")
            else:
                record.deposit_end_s = end_s
                record.completed_s = end_s
                self.completed_count += 1
                vehicle.carrier_id = None
                vehicle.current_task = None
                self._activate_next_task(vehicle, end_s)

    def _handle_arrivals(self, at_s: float) -> None:
        # 同一时刻到达同一 Port 时按载货、高优先级、车辆号确定唯一顺序。
        arrived: Dict[str, List[Vehicle]] = defaultdict(list)
        for vehicle in self.vehicles.values():
            if vehicle.state not in {"TO_PICKUP", "TO_DROPOFF"}:
                continue
            if self._remaining_distance(vehicle) <= 1e-5 and vehicle.speed <= 1e-5:
                assert vehicle.target_port is not None
                arrived[vehicle.target_port].append(vehicle)
        for port_id, vehicles in arrived.items():
            vehicles.sort(key=lambda v: (0 if v.state == "TO_DROPOFF" else 1, -self._task_priority(v), v.vehicle_id))
            for vehicle in vehicles:
                if port_id in self.port_owner:
                    vehicle.last_action = "停车"
                    vehicle.last_reason = "Port等待"
                    continue
                task_id = vehicle.current_task
                assert task_id is not None
                record = self.records[task_id]
                self.port_owner[port_id] = vehicle.vehicle_id
                vehicle.speed = 0.0
                vehicle.acceleration = 0.0
                if vehicle.state == "TO_PICKUP":
                    record.source_arrived_s = at_s
                    record.acquire_start_s = at_s
                    vehicle.state = "PICKING"
                    vehicle.service_remaining = PICK_TIME
                    vehicle.last_action = "取货"
                else:
                    record.destination_arrived_s = at_s
                    record.deposit_start_s = at_s
                    vehicle.state = "DROPPING"
                    vehicle.service_remaining = DROP_TIME
                    vehicle.last_action = "放货"
                vehicle.last_reason = "Port作业"

    def _update_pause(self, vehicle: Vehicle, reason: str) -> None:
        traffic_reasons = {"安全跟驰", "汇流避让", "弯轨等待", "Port等待"}
        active = vehicle.current_task is not None and vehicle.speed <= 1e-5 and reason in traffic_reasons
        if active:
            record = self.records[vehicle.current_task]
            record.pausing_ms += int(round(DT * 1000))
            if not vehicle.pause_active:
                record.paused_count += 1
        vehicle.pause_active = active
        vehicle.pause_reason = reason if active else ""

    def _violate(self, kind: str, vehicle_id: str, detail: str) -> None:
        self.violations.append({
            "time_s": round(self.time_s, 6), "type": kind,
            "vehicle_id": vehicle_id, "detail": detail,
        })

    def _audit_state(self, at_s: float) -> None:
        for vehicle in self.vehicles.values():
            link = self.data.links[vehicle.link_id]
            if not (-1e-5 <= vehicle.offset <= link.length + 1e-5):
                self._violate("position", vehicle.vehicle_id, f"{vehicle.offset} 不在 Link{link.link_id}")
            if not (-1e-5 <= vehicle.speed <= link.vmax + 1e-5):
                self._violate("speed", vehicle.vehicle_id, f"{vehicle.speed} > {link.vmax}")
            if not (-DEC - 1e-5 <= vehicle.acceleration <= ACC + 1e-5):
                self._violate("acceleration", vehicle.vehicle_id, str(vehicle.acceleration))
            self.max_observed_speed = max(self.max_observed_speed, vehicle.speed)
            self.max_observed_acc = max(self.max_observed_acc, vehicle.acceleration)
            self.min_observed_acc = min(self.min_observed_acc, vehicle.acceleration)

        by_link: Dict[int, List[Vehicle]] = defaultdict(list)
        for vehicle in self.vehicles.values():
            by_link[vehicle.link_id].append(vehicle)
        for link_id, vehicles in by_link.items():
            vehicles.sort(key=lambda v: v.offset)
            for back, front in zip(vehicles, vehicles[1:]):
                gap = front.offset - back.offset
                self.min_same_link_gap = min(self.min_same_link_gap, gap)
                if gap < REF_GAP - 1e-4:
                    self._violate("same_link_gap", back.vehicle_id, f"Link{link_id}, 前车{front.vehicle_id}, 间距{gap:.6f}")

        group_occupants: Dict[str, List[str]] = defaultdict(list)
        for vehicle in self.vehicles.values():
            group = self.data.links[vehicle.link_id].curve_group
            if group:
                group_occupants[group].append(vehicle.vehicle_id)
        for group, occupants in group_occupants.items():
            if len(occupants) > 1:
                self._violate("curve_capacity", ";".join(occupants), group)
            owner = self.curve_owner.get(group)
            if owner not in occupants and group not in self.curve_release:
                self._violate("curve_owner_missing", occupants[0], group)

        if len(set(self.port_owner.values())) != len(self.port_owner.values()):
            self._violate("vehicle_two_ports", "", str(self.port_owner))

    def _vehicle_state_label(self, vehicle: Vehicle) -> str:
        if vehicle.pause_active:
            return "等待"
        return {
            "IDLE": "待命", "TO_PICKUP": "前往取货", "PICKING": "取货",
            "TO_DROPOFF": "载货运输", "DROPPING": "放货",
        }.get(vehicle.state, vehicle.state)

    def _log_trajectory(self) -> None:
        for vehicle_id in sorted(self.vehicles):
            vehicle = self.vehicles[vehicle_id]
            next_seg = self._next_segment(vehicle) if vehicle.state in self.MOVING_STATES else None
            self.trajectory.append([
                1, self.step_no, round(self.time_s, 3), vehicle.vehicle_id,
                vehicle.current_task or "", vehicle.carrier_id or "",
                self._vehicle_state_label(vehicle), vehicle.link_id, round(vehicle.offset, 3),
                round(vehicle.speed, 3), next_seg.link_id if next_seg else "",
                vehicle.last_action, vehicle.last_reason, vehicle.related_vehicle,
            ])

    def run(self, max_time_s: float = 6000.0) -> Dict[str, object]:
        started = time.perf_counter()
        while self.completed_count < len(self.data.tasks):
            if self.time_s >= max_time_s - EPS:
                self._violate("timeout", "", f"{max_time_s}s 未完成全部任务")
                break
            self._arbitrate_resources()
            end_s = round(self.time_s + DT, 10)
            self._process_services(end_s)

            controls: Dict[str, Tuple[float, float, str, str, float, bool]] = {}
            for vehicle in self.vehicles.values():
                if vehicle.state not in self.MOVING_STATES:
                    continue
                old_speed = vehicle.speed
                boundary_allowed = not self._boundary_block(vehicle)[0]
                movement, new_speed, reason, related = self._movement_control(vehicle)
                controls[vehicle.vehicle_id] = (movement, new_speed, reason, related, old_speed, boundary_allowed)

            for vehicle_id, (movement, new_speed, reason, related, old_speed, boundary_allowed) in controls.items():
                vehicle = self.vehicles[vehicle_id]
                self._advance_vehicle(vehicle, movement, new_speed, boundary_allowed)
                vehicle.acceleration = (vehicle.speed - old_speed) / DT
                vehicle.last_action = "停车" if vehicle.speed <= 1e-5 else ("减速" if vehicle.acceleration < -1e-5 else "运行")
                vehicle.last_reason = reason
                vehicle.related_vehicle = related
                self._update_pause(vehicle, reason)

            self.time_s = end_s
            self.step_no += 1
            self._handle_arrivals(self.time_s)
            self._audit_state(self.time_s)
            self._log_trajectory()

        runtime = time.perf_counter() - started
        completed = [r for r in self.records.values() if r.completed_s is not None]
        result = {
            "label": self.label,
            "completed_tasks": len(completed),
            "makespan_s": max((r.completed_s or 0.0) for r in completed) if completed else math.inf,
            "avg_transfer_time_s": statistics.mean(r.completed_s for r in completed) if completed else math.inf,
            "simulation_runtime_s": runtime,
            "hard_violation_count": len(self.violations),
            "trajectory_rows": len(self.trajectory),
            "min_same_link_reference_gap_mm": None if math.isinf(self.min_same_link_gap) else self.min_same_link_gap,
            "max_speed_mm_s": self.max_observed_speed,
            "max_acc_mm_s2": self.max_observed_acc,
            "min_acc_mm_s2": self.min_observed_acc,
        }
        return result


TASK_HEADERS = [
    "QuestionNo", "CommandID", "CarrierID", "Priority", "InstallTime", "Source", "Destination",
    "VehicleID", "AssignedTime", "WaitAssignTime", "VehicleFromArrivedTime",
    "VehicleAcquireStartTime", "VehicleAcquireEndTime", "VehicleDepartedTime",
    "VehicleToArrivedTime", "VehicleDepositStartTime", "VehicleDepositEndTime",
    "TransferCompletedTime", "ToSourceDistance", "ToDestinationDistance", "TransferTime",
    "MoveToSourcePath", "MoveToDestinationPath", "PausedCount", "PausingTime",
]

TRAJECTORY_HEADERS = [
    "QuestionNo", "StepNo", "SimTime", "VehicleID", "CommandID", "CarrierID", "VehicleState",
    "CurrentEdgeID", "Position", "Speed", "NextEdgeID", "Action", "Reason", "RelatedVehicleID",
]

METRIC_HEADERS = [
    "QuestionNo", "TaskCount", "AvgTransferTime", "AvgAssignTime", "AvgPickupResponseTime",
    "AvgTransportTime", "AvgPausingTime", "Makespan", "AvgEmptyDistance", "AvgTotalDistance",
    "SimulationRuntime",
]


def absolute_time(base: datetime, relative_s: Optional[float]) -> Optional[datetime]:
    if relative_s is None:
        return None
    return base + timedelta(milliseconds=round(relative_s * 1000.0))


def task_rows(sim: TrafficSimulator) -> List[List[object]]:
    base = min(task.install_time for task in sim.data.tasks.values())
    rows: List[List[object]] = []
    for task in sorted(sim.data.tasks.values(), key=lambda x: x.command_id):
        record = sim.records[task.command_id]
        if record.completed_s is None:
            raise RuntimeError(f"任务未完成，不能导出: {task.command_id}")
        rows.append([
            1, task.command_id, task.carrier_id, task.priority, task.install_time, task.source, task.destination,
            record.vehicle_id, absolute_time(base, record.assigned_s), int(round(record.assigned_s * 1000)),
            absolute_time(base, record.source_arrived_s), absolute_time(base, record.acquire_start_s),
            absolute_time(base, record.acquire_end_s), absolute_time(base, record.departed_s),
            absolute_time(base, record.destination_arrived_s), absolute_time(base, record.deposit_start_s),
            absolute_time(base, record.deposit_end_s), absolute_time(base, record.completed_s),
            round(record.to_source_distance, 3), round(record.to_destination_distance, 3),
            round(record.completed_s, 3), record.move_to_source_path, record.move_to_destination_path,
            record.paused_count, record.pausing_ms,
        ])
    return rows


def metric_row(sim: TrafficSimulator, runtime_s: float) -> List[object]:
    records = list(sim.records.values())
    if any(r.completed_s is None for r in records):
        raise RuntimeError("存在未完成任务，不能计算指标")
    return [
        1,
        len(records),
        round(statistics.mean(r.completed_s for r in records), 6),
        round(statistics.mean(r.assigned_s for r in records), 6),
        round(statistics.mean(r.source_arrived_s - r.assigned_s for r in records), 6),
        round(statistics.mean(r.destination_arrived_s - r.departed_s for r in records), 6),
        round(statistics.mean(r.pausing_ms / 1000.0 for r in records), 6),
        round(max(r.completed_s for r in records), 6),
        round(statistics.mean(r.to_source_distance for r in records), 6),
        round(statistics.mean(r.to_source_distance + r.to_destination_distance for r in records), 6),
        round(runtime_s, 6),
    ]


def write_csv(path: Path, headers: Sequence[str], rows: Iterable[Sequence[object]]) -> None:
    with path.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.writer(handle)
        writer.writerow(headers)
        writer.writerows(rows)


def export_excel(root: Path, out_path: Path, tasks: List[List[object]], trajectory: List[List[object]], metrics: List[object]) -> None:
    template = root / "附件9_结果提交模板.xlsx"
    wb = openpyxl.load_workbook(template)
    ws_task = wb["任务仿真结果"]
    ws_trace = wb["OHT逐步运行记录表"]
    ws_metric = wb["算法评价指标"]
    if [cell.value for cell in ws_task[1]] != TASK_HEADERS:
        raise ValueError("附件9任务结果表头发生变化")
    if [cell.value for cell in ws_trace[1]] != TRAJECTORY_HEADERS:
        raise ValueError("附件9轨迹表头发生变化")
    if [cell.value for cell in ws_metric[1]] != METRIC_HEADERS:
        raise ValueError("附件9指标表头发生变化")
    for row_idx, values in enumerate(tasks, start=2):
        for col_idx, value in enumerate(values, start=1):
            cell = ws_task.cell(row=row_idx, column=col_idx, value=value)
            if isinstance(value, datetime):
                cell.number_format = "yyyy-mm-dd hh:mm:ss.000"
    for values in trajectory:
        ws_trace.append(values)
    for col_idx, value in enumerate(metrics, start=1):
        ws_metric.cell(row=2, column=col_idx, value=value)
    wb.save(out_path)
    # 保存后重读，防止产生不可读工作簿或写错行数。
    check = openpyxl.load_workbook(out_path, data_only=False, read_only=True)
    if check["任务仿真结果"].max_row != 823:
        raise ValueError("结果工作簿任务模板行数被破坏")
    if check["OHT逐步运行记录表"].max_row != len(trajectory) + 1:
        raise ValueError("结果工作簿轨迹行数不一致")
    if check["算法评价指标"].cell(2, 2).value != 32:
        raise ValueError("结果工作簿问题1任务数未写入")


def final_audit(data: DataBundle, sim: TrafficSimulator, summary: Dict[str, object]) -> Dict[str, object]:
    task_checks: List[str] = []
    for task_id, record in sim.records.items():
        times = [
            record.assigned_s, record.source_arrived_s, record.acquire_start_s, record.acquire_end_s,
            record.departed_s, record.destination_arrived_s, record.deposit_start_s,
            record.deposit_end_s, record.completed_s,
        ]
        if any(x is None for x in times):
            task_checks.append(f"{task_id}: 时间字段不完整")
            continue
        if any(a > b + EPS for a, b in zip(times, times[1:])):
            task_checks.append(f"{task_id}: 时间字段非单调")
        if abs((record.acquire_end_s - record.acquire_start_s) - PICK_TIME) > 1e-6:
            task_checks.append(f"{task_id}: 取货服务时间不等于 8s")
        if abs((record.deposit_end_s - record.deposit_start_s) - DROP_TIME) > 1e-6:
            task_checks.append(f"{task_id}: 放货服务时间不等于 8s")
        if not record.move_to_source_path or not record.move_to_destination_path:
            task_checks.append(f"{task_id}: 路径字段为空")
    step_counts: Dict[int, int] = defaultdict(int)
    for row in sim.trajectory:
        step_counts[int(row[1])] += 1
    bad_steps = [step for step, count in step_counts.items() if count != 20]
    output = {
        "data_audit": data.audit,
        "simulation_summary": summary,
        "completed_task_ids": sorted(x for x, r in sim.records.items() if r.completed_s is not None),
        "task_validation_errors": task_checks,
        "bad_trajectory_steps": bad_steps,
        "traffic_violations": sim.violations,
    }
    output["passed"] = (
        data.audit["passed"]
        and summary["completed_tasks"] == 32
        and summary["hard_violation_count"] == 0
        and not task_checks
        and not bad_steps
    )
    return output


def write_result_note(path: Path, selected_label: str, baseline: Dict[str, object], insertion: Dict[str, object], improved: Dict[str, object], metrics: List[object], audit: Dict[str, object]) -> None:
    text = f"""# 问题1结果说明

## 完成情况

- 最终采用方案：{selected_label}
- 完成任务数：{audit['simulation_summary']['completed_tasks']}/32
- 硬约束违规数：{audit['simulation_summary']['hard_violation_count']}
- 数据与轨迹总审计：{'通过' if audit['passed'] else '未通过'}

## 两套方案真实仿真比较

| 方案 | 平均任务执行时间/s | Makespan/s | 硬违规数 |
|---|---:|---:|---:|
| 基准追加调度 | {baseline['avg_transfer_time_s']:.3f} | {baseline['makespan_s']:.3f} | {baseline['hard_violation_count']} |
| 全局增量插入 | {insertion['avg_transfer_time_s']:.3f} | {insertion['makespan_s']:.3f} | {insertion['hard_violation_count']} |
| ALNS静态调度 | {improved['avg_transfer_time_s']:.3f} | {improved['makespan_s']:.3f} | {improved['hard_violation_count']} |

## 附件9指标

| 指标 | 数值 |
|---|---:|
| AvgTransferTime/s | {metrics[2]} |
| AvgAssignTime/s | {metrics[3]} |
| AvgPickupResponseTime/s | {metrics[4]} |
| AvgTransportTime/s | {metrics[5]} |
| AvgPausingTime/s | {metrics[6]} |
| Makespan/s | {metrics[7]} |
| AvgEmptyDistance/mm | {metrics[8]} |
| AvgTotalDistance/mm | {metrics[9]} |
| SimulationRuntime/s | {metrics[10]} |

## 模型与实现口径

第一问采用带资源约束的静态多车取送路径模型。高层比较优先级 FCFS 最近完成车辆基准与代理成本 ALNS；底层统一采用位置感知有向最短路、0.2 s 同步运动控制、Port 服务、MERGE/CURVE 仲裁和跟驰控制。

P17 在读取后、构图前修订为 `(Node70, Link507, 303 mm)`。安全约束按车身净空 300 mm 实现；由于车辆长度为 950 mm，同参考点最小间距为 1250 mm。

## 复现

在当前目录运行：

```powershell
python .\\q1_solver.py
```

程序会重新读取附件、运行两套方案、选择零违规且平均任务执行时间较小的方案，并覆盖生成 `outputs/q1` 下的正式结果。
"""
    path.write_text(text, encoding="utf-8")


def solve(root: Path, iterations: int) -> Dict[str, object]:
    output_dir = root / "outputs" / "q1"
    output_dir.mkdir(parents=True, exist_ok=True)
    data = DataBundle(root).load()
    graph = GraphEngine(data)
    scheduler = StaticScheduler(data, graph)

    baseline_chains = scheduler.baseline()
    insertion_chains = scheduler.insertion_initial()
    improved_chains, alns_details = scheduler.alns(iterations=iterations)

    print("运行问题1基准方案物理仿真……", flush=True)
    baseline_sim = TrafficSimulator(data, graph, baseline_chains, "baseline")
    baseline_summary = baseline_sim.run()
    print(json.dumps(baseline_summary, ensure_ascii=False, indent=2), flush=True)

    print("运行问题1全局增量插入方案物理仿真……", flush=True)
    insertion_sim = TrafficSimulator(data, graph, insertion_chains, "insertion")
    insertion_summary = insertion_sim.run()
    print(json.dumps(insertion_summary, ensure_ascii=False, indent=2), flush=True)

    print("运行问题1 ALNS 方案物理仿真……", flush=True)
    improved_sim = TrafficSimulator(data, graph, improved_chains, "ALNS")
    improved_summary = improved_sim.run()
    print(json.dumps(improved_summary, ensure_ascii=False, indent=2), flush=True)

    candidates = [
        (baseline_summary, baseline_sim, baseline_chains, "基准追加调度"),
        (insertion_summary, insertion_sim, insertion_chains, "全局增量插入"),
        (improved_summary, improved_sim, improved_chains, "ALNS静态调度"),
    ]
    feasible = [x for x in candidates if x[0]["completed_tasks"] == 32 and x[0]["hard_violation_count"] == 0]
    if not feasible:
        diagnostics = {
            "baseline": baseline_summary,
            "baseline_violations": baseline_sim.violations[:100],
            "insertion": insertion_summary,
            "insertion_violations": insertion_sim.violations[:100],
            "improved": improved_summary,
            "improved_violations": improved_sim.violations[:100],
        }
        (output_dir / "问题1_失败诊断.json").write_text(json.dumps(diagnostics, ensure_ascii=False, indent=2), encoding="utf-8")
        raise RuntimeError("三套方案均未通过硬约束审计，已写出失败诊断")
    stale_failure = output_dir / "问题1_失败诊断.json"
    if stale_failure.exists():
        stale_failure.unlink()
    selected_summary, selected_sim, selected_chains, selected_label = min(
        feasible, key=lambda x: (x[0]["avg_transfer_time_s"], x[0]["makespan_s"])
    )
    rows = task_rows(selected_sim)
    metrics = metric_row(selected_sim, selected_summary["simulation_runtime_s"])
    audit = final_audit(data, selected_sim, selected_summary)
    if not audit["passed"]:
        raise RuntimeError("最终方案未通过完整审计")

    write_csv(output_dir / "问题1_任务结果.csv", TASK_HEADERS, rows)
    write_csv(output_dir / "问题1_OHT逐步运行记录.csv", TRAJECTORY_HEADERS, selected_sim.trajectory)
    write_csv(output_dir / "问题1_算法评价指标.csv", METRIC_HEADERS, [metrics])
    (output_dir / "问题1_约束审计.json").write_text(json.dumps(audit, ensure_ascii=False, indent=2), encoding="utf-8")
    plan_output = {
        "selected_label": selected_label,
        "selected_chains": selected_chains,
        "baseline_chains": baseline_chains,
        "insertion_chains": insertion_chains,
        "alns_chains": improved_chains,
        "alns_details": alns_details,
        "baseline_summary": baseline_summary,
        "insertion_summary": insertion_summary,
        "improved_summary": improved_summary,
        "configuration": {
            "dt_s": DT, "acc_mm_s2": ACC, "dec_mm_s2": DEC,
            "vehicle_length_mm": VEHICLE_LENGTH, "clear_gap_mm": CLEAR_GAP,
            "reference_gap_mm": REF_GAP, "pick_time_s": PICK_TIME, "drop_time_s": DROP_TIME,
            "p17_override": [70, 507, 303.0],
        },
    }
    (output_dir / "问题1_方案.json").write_text(json.dumps(plan_output, ensure_ascii=False, indent=2), encoding="utf-8")
    export_excel(root, output_dir / "问题1_结果.xlsx", rows, selected_sim.trajectory, metrics)
    write_result_note(output_dir / "问题1_结果说明.md", selected_label, baseline_summary, insertion_summary, improved_summary, metrics, audit)
    return {
        "selected": selected_label,
        "metrics": dict(zip(METRIC_HEADERS, metrics)),
        "audit_passed": audit["passed"],
        "output_dir": str(output_dir),
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="求解 OHT 赛题问题1")
    parser.add_argument("--root", type=Path, default=Path(__file__).resolve().parent)
    parser.add_argument("--iterations", type=int, default=1200)
    args = parser.parse_args()
    result = solve(args.root.resolve(), args.iterations)
    print("问题1完成：", flush=True)
    print(json.dumps(result, ensure_ascii=False, indent=2), flush=True)


if __name__ == "__main__":
    main()
