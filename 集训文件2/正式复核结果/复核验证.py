from __future__ import annotations

import csv
import hashlib
import json
import sys
from collections import Counter
from pathlib import Path

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "outputs"
EVIDENCE = Path(__file__).with_name("统一验证证据.json")
COUNTERFACTUAL = Path(__file__).with_name("在线反事实验证.json")
CONTINUOUS = Path(__file__).with_name("独立连续净空复核证据.json")
EXPECTED_TASKS = {1: 32, 2: 190, 3: 600}


def csv_rows(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def verify_question(question: int) -> dict[str, object]:
    folder = OUT / f"q{question}"
    prefix = f"问题{question}"
    task_path = folder / f"{prefix}_任务结果.csv"
    trace_path = folder / f"{prefix}_OHT逐步运行记录.csv"
    metric_path = folder / f"{prefix}_算法评价指标.csv"
    audit_path = folder / f"{prefix}_约束审计.json"
    plan_path = folder / f"{prefix}_方案.json"
    xlsx_path = folder / f"{prefix}_结果.xlsx"
    tasks = csv_rows(task_path)
    trace = csv_rows(trace_path)
    metrics = csv_rows(metric_path)
    audit = json.loads(audit_path.read_text(encoding="utf-8"))
    plan = json.loads(plan_path.read_text(encoding="utf-8"))
    errors: list[str] = []
    expected = EXPECTED_TASKS[question]

    if len(tasks) != expected or len({row["CommandID"] for row in tasks}) != expected:
        errors.append("任务数量或唯一性错误")
    if any(int(row["QuestionNo"]) != question for row in tasks + trace + metrics):
        errors.append("CSV QuestionNo 错误")
    steps = Counter(int(row["StepNo"]) for row in trace)
    if not steps or set(steps.values()) != {20} or sorted(steps) != list(range(max(steps) + 1)):
        errors.append("轨迹步不连续或每步不是20车")
    summary = audit.get("simulation_summary", {})
    if not audit.get("passed") or audit.get("traffic_violations"):
        errors.append("内部约束审计未通过")
    clearance = summary.get("min_continuous_clearance_mm")
    if clearance is None or float(clearance) < 300.0 - 1e-4:
        errors.append("步内连续最小净空不足300mm")
    if int(summary.get("completed_tasks", -1)) != expected:
        errors.append("任务未全部完成")
    if question >= 2:
        if audit.get("task_validation_errors") or audit.get("online_validation_errors"):
            errors.append("Carrier或在线性审计失败")
        if int(summary.get("future_leak_count", -1)) != 0:
            errors.append("存在未来信息泄漏")
    if question == 3 and audit.get("slot_audit", {}).get("actual_overlap_errors"):
        errors.append("资源实际占用重叠")

    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    trace_sheet = wb["OHT逐步运行记录表"]
    if trace_sheet.max_row != len(trace) + 1:
        errors.append("Excel与CSV轨迹行数不一致")
    if any(int(row[0].value) != question for row in trace_sheet.iter_rows(min_row=2, max_col=1)):
        errors.append("Excel轨迹QuestionNo错误")
    metric_row = question + 1
    if wb["算法评价指标"].cell(metric_row, 1).value != question:
        errors.append("Excel指标QuestionNo错误")

    if question == 1:
        candidates = {
            "baseline": plan["baseline_summary"],
            "insertion": plan["insertion_summary"],
            "ALNS": plan["improved_summary"],
        }
        selected_metric = float(metrics[0]["AvgTransferTime"])
        feasible_values = [
            float(item["avg_transfer_time_s"]) for item in candidates.values()
            if item.get("hard_gate_passed") is True
        ]
    else:
        candidates = plan["summaries"]
        selected = plan["selected_mode"]
        selected_metric = float(candidates[selected]["avg_transfer_time_s"])
        feasible_values = [
            float(item["avg_transfer_time_s"]) for item in candidates.values()
            if item["completed_tasks"] == expected
            and item["hard_violation_count"] == 0
            and item["future_leak_count"] == 0
            and float(item["min_continuous_clearance_mm"]) >= 300.0 - 1e-4
        ]
    if not feasible_values or abs(selected_metric - min(feasible_values)) > 1e-6:
        errors.append("正式方案不是硬门通过候选中的AvgTransferTime最优")

    paths = [task_path, trace_path, metric_path, audit_path, plan_path, xlsx_path]
    return {
        "question": question,
        "passed": not errors,
        "errors": errors,
        "task_rows": len(tasks),
        "trajectory_rows": len(trace),
        "steps": len(steps),
        "completed_tasks": summary.get("completed_tasks"),
        "min_continuous_clearance_mm": clearance,
        "hard_violation_count": summary.get("hard_violation_count"),
        "future_leak_count": summary.get("future_leak_count", 0),
        "selected_avg_transfer_time_s": selected_metric,
        "file_sha256": {str(path.relative_to(ROOT)): sha256(path) for path in paths},
    }


def main() -> None:
    results = [verify_question(question) for question in (1, 2, 3)]
    counterfactual = json.loads(COUNTERFACTUAL.read_text(encoding="utf-8"))
    continuous = json.loads(CONTINUOUS.read_text(encoding="utf-8"))
    evidence = {
        "passed": (
            all(item["passed"] for item in results)
            and counterfactual.get("passed") is True
            and continuous.get("passed") is True
        ),
        "hard_gate": {
            "minimum_clearance_mm": 300.0,
            "all_tasks_completed": True,
            "resource_carrier_online_required": True,
        },
        "questions": results,
        "online_counterfactual": counterfactual,
        "independent_continuous_clearance": continuous,
        "python": sys.version,
    }
    EVIDENCE.write_text(json.dumps(evidence, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(evidence, ensure_ascii=False, indent=2))
    if not evidence["passed"]:
        raise SystemExit(1)


if __name__ == "__main__":
    main()
