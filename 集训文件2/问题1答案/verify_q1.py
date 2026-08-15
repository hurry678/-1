from __future__ import annotations

import csv
import json
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path

import openpyxl


ROOT = Path(__file__).resolve().parent
OUT = ROOT / "outputs" / "q1"
DT = 0.2
ACC = 2000.0
DEC = 3000.0
REF_GAP = 1250.0


def read_csv(name: str):
    with (OUT / name).open("r", encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def main() -> None:
    errors: list[str] = []
    link_rows = openpyxl.load_workbook(
        ROOT / "附件2_轨道连接数据.xlsx", data_only=True, read_only=True
    )["Link"].iter_rows(min_row=2, values_only=True)
    links = {int(r[0]): (float(r[4]), float(r[5])) for r in link_rows if int(r[1]) == 1}

    tasks = read_csv("问题1_任务结果.csv")
    trace = read_csv("问题1_OHT逐步运行记录.csv")
    metrics = read_csv("问题1_算法评价指标.csv")
    if len(tasks) != 32:
        errors.append(f"任务结果行数={len(tasks)}")
    if len({r["CommandID"] for r in tasks}) != 32:
        errors.append("CommandID 不唯一")

    time_fmt = "%Y-%m-%d %H:%M:%S.%f"
    for row in tasks:
        values = [
            datetime.strptime(row[k], time_fmt)
            for k in [
                "InstallTime", "AssignedTime", "VehicleFromArrivedTime", "VehicleAcquireStartTime",
                "VehicleAcquireEndTime", "VehicleDepartedTime", "VehicleToArrivedTime",
                "VehicleDepositStartTime", "VehicleDepositEndTime", "TransferCompletedTime",
            ]
        ]
        if any(a > b for a, b in zip(values, values[1:])):
            errors.append(f"{row['CommandID']} 时间非单调")
        if abs((values[4] - values[3]).total_seconds() - 8.0) > 1e-9:
            errors.append(f"{row['CommandID']} 取货时间错误")
        if abs((values[8] - values[7]).total_seconds() - 8.0) > 1e-9:
            errors.append(f"{row['CommandID']} 放货时间错误")
        transfer = (values[9] - values[0]).total_seconds()
        if abs(transfer - float(row["TransferTime"])) > 1e-6:
            errors.append(f"{row['CommandID']} TransferTime 错误")

    step_counts = Counter(int(r["StepNo"]) for r in trace)
    bad_steps = [s for s, count in step_counts.items() if count != 20]
    if bad_steps:
        errors.append(f"非20车仿真步: {bad_steps[:10]}")
    expected_steps = list(range(max(step_counts) + 1))
    if sorted(step_counts) != expected_steps:
        errors.append("StepNo 不连续")

    previous_speed: dict[str, float] = {}
    for row in trace:
        link_id = int(row["CurrentEdgeID"])
        position = float(row["Position"])
        speed = float(row["Speed"])
        length, vmax = links[link_id]
        if not (-1e-6 <= position <= length + 1e-6):
            errors.append(f"{row['VehicleID']} 位置越界")
        if not (-1e-6 <= speed <= vmax + 1e-6):
            errors.append(f"{row['VehicleID']} 速度越界")
        vehicle_id = row["VehicleID"]
        if vehicle_id in previous_speed:
            acceleration = (speed - previous_speed[vehicle_id]) / DT
            if acceleration < -DEC - 1e-6 or acceleration > ACC + 1e-6:
                errors.append(f"{vehicle_id} 加速度越界 {acceleration}")
        previous_speed[vehicle_id] = speed

    by_step_link: dict[tuple[int, int], list[tuple[float, str]]] = defaultdict(list)
    for row in trace:
        by_step_link[(int(row["StepNo"]), int(row["CurrentEdgeID"]))].append(
            (float(row["Position"]), row["VehicleID"])
        )
    min_gap = float("inf")
    for (step, link_id), vehicles in by_step_link.items():
        vehicles.sort()
        for back, front in zip(vehicles, vehicles[1:]):
            gap = front[0] - back[0]
            min_gap = min(min_gap, gap)
            if gap < REF_GAP - 1e-6:
                errors.append(f"Step{step} Link{link_id} 间距={gap}")

    wb = openpyxl.load_workbook(OUT / "问题1_结果.xlsx", data_only=True, read_only=True)
    if wb["任务仿真结果"].max_row != 823:
        errors.append("Excel任务模板行数错误")
    if wb["OHT逐步运行记录表"].max_row != len(trace) + 1:
        errors.append("Excel轨迹行数与CSV不一致")
    if wb["算法评价指标"].cell(2, 2).value != 32:
        errors.append("Excel指标任务数错误")

    result = {
        "passed": not errors,
        "errors": errors[:100],
        "task_rows": len(tasks),
        "trajectory_rows": len(trace),
        "steps": len(step_counts),
        "vehicles_per_step": sorted(set(step_counts.values())),
        "min_same_link_reference_gap_mm": min_gap,
        "metric_row": metrics[0] if metrics else None,
    }
    print(json.dumps(result, ensure_ascii=False, indent=2))
    if errors:
        raise SystemExit(1)


if __name__ == "__main__":
    main()
