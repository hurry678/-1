from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "三问全过程计划与答案二次复核汇总.xlsx"

NAVY = "17365D"
BLUE = "DCEAF7"
LIGHT = "F7F9FC"
GREEN = "E2F0D9"
YELLOW = "FFF2CC"
RED = "FCE4D6"
WHITE = "FFFFFF"
GRID = Side(style="thin", color="D9DEE7")


def style_sheet(ws, widths):
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for cell in ws[1]:
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.font = Font(name="Arial", bold=True, color=WHITE)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in ws.iter_rows(min_row=2):
        fill = PatternFill("solid", fgColor=WHITE if row[0].row % 2 == 0 else LIGHT)
        for cell in row:
            cell.font = Font(name="Arial", size=10)
            cell.fill = fill
            cell.border = Border(bottom=GRID)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    ws.row_dimensions[1].height = 30


def fill_status(cell, status):
    color = GREEN if status in {"通过", "有条件通过", "正确"} else RED if status in {"不通过", "错误"} else YELLOW
    cell.fill = PatternFill("solid", fgColor=color)
    cell.font = Font(name="Arial", bold=True, color="1F2937")


wb = Workbook()
ws = wb.active
ws.title = "总览"
ws.append([
    "问题", "正式方案", "完成任务", "AvgTransferTime/s",
    "程序内连续净空/mm", "采样路径净空/mm", "安全裕量/mm",
    "制动距离违规数", "版本一致", "最终判定", "说明",
])
summary = [
    [1, "ALNS", "32/32", 171.5, 325.0, 325.0, "=F2-300", 0, "是", "通过", "修复后的分段连续审计与独立停车距离审计均通过"],
    [2, "balanced", "190/190", 142.939989, 300.5, 300.5, "=F3-300", 0, "是", "通过", "原3个停车违规已修复，四个候选均通过硬门"],
    [3, "microbatch", "600/600", 830.458087, 307.8, 307.8, "=F4-300", 0, "是", "通过", "公共内核回归通过；完整时隙模型保留为候选"],
]
for row in summary:
    ws.append(row)
style_sheet(ws, {"A": 8, "B": 18, "C": 12, "D": 20, "E": 21, "F": 19, "G": 14, "H": 16, "I": 12, "J": 14, "K": 52})
for r in range(2, 5):
    fill_status(ws.cell(r, 10), ws.cell(r, 10).value)
ws.conditional_formatting.add("G2:G4", CellIsRule(operator="lessThan", formula=["0"], fill=PatternFill("solid", fgColor=RED)))
ws.conditional_formatting.add("H2:H4", CellIsRule(operator="greaterThan", formula=["0"], fill=PatternFill("solid", fgColor=RED)))

ws = wb.create_sheet("计划核对")
ws.append(["类别", "核对项", "判定", "证据/问题", "建议修改"])
plan_rows = [
    ["安全口径", "300 mm净空等价于1250 mm同参考点间距", "正确", "题目、计划与代码常量一致", "保留"],
    ["连续安全", "跨Link并检查0.2 s步内极值", "正确", "已按实际常加速度/匀速/最大制动/静止分段逐段求极值", "保留分段回归测试"],
    ["运动学", "速度、加减速与停车距离", "正确", "Q2原3个停车违规已清零，三问独立停车距离违规均为0", "保留v²/(2DEC)独立硬门"],
    ["P17", "Node70、Link507、距起点303 mm", "正确", "附件2/3与代码断言一致", "保留并清理旧网页缓存"],
    ["在线性", "Q2/Q3不得读取未来任务", "正确", "6组未来字段扰动哈希一致", "保留反事实测试"],
    ["主目标", "可行候选中AvgTransferTime最小", "正确", "Q1/Q2/Q3分别选择ALNS、balanced、microbatch", "保留"],
    ["Q3模型", "压力+时隙+防死锁作为正式主模型", "需修订", "正式轨迹为microbatch，时隙/预约/恢复事件均为0", "调整模型定位或优化完整模型"],
    ["附件9行数", "三问任务数据总行数", "需修订", "32+190+600=822；823只能表示含表头", "统一措辞"],
    ["版本管理", "README/报告/CSV/XLSX同版本", "正确", "正式输出已同步回三个答案目录", "提交前继续执行哈希复核"],
]
for row in plan_rows:
    ws.append(row)
style_sheet(ws, {"A": 14, "B": 34, "C": 14, "D": 58, "E": 44})
for r in range(2, ws.max_row + 1):
    fill_status(ws.cell(r, 3), ws.cell(r, 3).value)

ws = wb.create_sheet("运动学反例")
ws.append([
    "问题", "修复前Step", "车辆", "修复前初速度/mm·s⁻¹", "修复前实际位移/mm",
    "最短停车距离/mm", "修复前缺口/mm", "修复后判定", "说明",
])
cases = [
    [2, 7264, "OHT08", 597.329, 51.385, "=D2^2/(2*3000)", "=F2-E2", "通过", "修复后正式轨迹未出现停车距离不足"],
    [2, 10506, "OHT09", 366.333, 14.285, "=D3^2/(2*3000)", "=F3-E3", "通过", "修复后正式轨迹未出现停车距离不足"],
    [2, 16891, "OHT02", 110.0, 0.668, "=D4^2/(2*3000)", "=F4-E4", "通过", "增加高速空闲车到近距离Source的制动接纳检查"],
]
for row in cases:
    ws.append(row)
style_sheet(ws, {"A": 10, "B": 12, "C": 14, "D": 20, "E": 18, "F": 22, "G": 14, "H": 14, "I": 16})
for r in range(2, 5):
    fill_status(ws.cell(r, 8), ws.cell(r, 8).value)
    for c in range(4, 8):
        ws.cell(r, c).number_format = "0.000000"

ws = wb.create_sheet("版本差异")
ws.append([
    "问题", "答案目录AvgTransferTime/s", "outputs正式AvgTransferTime/s",
    "差值/s", "答案目录状态", "应采用来源", "处置",
])
versions = [
    [1, 171.500000, 171.500000, "=C2-B2", "已同步", "outputs/q1", "一致"],
    [2, 142.939989, 142.939989, "=C3-B3", "已同步", "outputs/q2", "一致"],
    [3, 830.458087, 830.458087, "=C4-B4", "已同步", "outputs/q3", "一致"],
]
for row in versions:
    ws.append(row)
style_sheet(ws, {"A": 10, "B": 26, "C": 26, "D": 14, "E": 18, "F": 18, "G": 40})
for r in range(2, 5):
    ws.cell(r, 5).fill = PatternFill("solid", fgColor=GREEN)
    ws.cell(r, 5).font = Font(name="Arial", bold=True)

for ws in wb.worksheets:
    ws.sheet_view.showGridLines = False
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, (int, float)) and cell.column > 1:
                cell.number_format = "0.000000"

wb.calculation.fullCalcOnLoad = True
wb.calculation.forceFullCalc = True
wb.calculation.calcMode = "auto"
wb.save(OUT)
print(OUT)
