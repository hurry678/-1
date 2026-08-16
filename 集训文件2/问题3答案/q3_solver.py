from __future__ import annotations

import argparse
import json
import math
import statistics
import time
from collections import Counter, defaultdict
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence, Tuple

import openpyxl

import q1_solver as core
import q2_solver as online


MAX_TIME_S = 9000.0
HIGH_PRIORITY = 90
FULL_BATCH_S = 5.0
FULL_BATCH_COUNT = 6
SLOT_GAP_S = 0.2


def load_q3_data(root: Path) -> core.DataBundle:
    data = core.DataBundle(root).load()
    ws = openpyxl.load_workbook(
        root / "附件7_高密任务数据.xlsx", data_only=True, read_only=True
    )["Task"]
    tasks: Dict[str, core.Task] = {}
    for command_id, install_time, source, destination, priority, carrier_id in ws.iter_rows(
        min_row=2, values_only=True
    ):
        task = core.Task(
            str(command_id), install_time, str(source), str(destination),
            int(priority), str(carrier_id),
        )
        tasks[task.command_id] = task
    if len(tasks) != 600:
        raise ValueError(f"问题3任务数量 {len(tasks)} != 600")
    for task in tasks.values():
        if task.source not in data.ports or task.destination not in data.ports:
            raise ValueError(f"{task.command_id} Port外键无效")
    data.tasks = tasks
    carrier_count = Counter(t.carrier_id for t in tasks.values())
    base = min(t.install_time for t in tasks.values())
    data.audit = {
        **data.audit,
        "counts": {**data.audit["counts"], "tasks": 600},
        "scenario": "问题3",
        "release_span_s": (max(t.install_time for t in tasks.values()) - base).total_seconds(),
        "different_carriers": len(carrier_count),
        "repeated_carriers": sum(v > 1 for v in carrier_count.values()),
        "max_carrier_chain": max(carrier_count.values()),
        "errors": [],
        "passed": True,
    }
    return data


class HighDensityScheduler(online.OnlineScheduler):
    def __init__(self, data: core.DataBundle, graph: core.GraphEngine, q3_mode: str):
        base_mode = "rolling" if q3_mode == "direct" else "balanced"
        super().__init__(data, graph, base_mode)
        self.q3_mode = q3_mode

    def _best_pressure_insertion(
        self, sim: "HighDensitySimulator", task_id: str, now_s: float
    ) -> Tuple[str, int, float]:
        task = self.data.tasks[task_id]
        best: Optional[Tuple[float, str, int]] = None
        port_pressure = self._port_pressure(sim, task)
        region_pressure = sim.region_pressure_snapshot()
        source_region = sim.port_region(task.source)
        destination_region = sim.port_region(task.destination)
        mean_pressure = statistics.mean(region_pressure.values()) if region_pressure else 0.0
        marginal_pressure = (
            0.5 * max(0.0, region_pressure.get(source_region, 0.0) - mean_pressure)
            + max(0.0, region_pressure.get(destination_region, 0.0) - mean_pressure)
        )
        for vehicle_id in sorted(sim.vehicles):
            vehicle = sim.vehicles[vehicle_id]
            base_comp, base_empty, _ = self._forecast_chain(sim, vehicle, vehicle.queue, now_s)
            base_sum = sum(base_comp)
            for pos in range(len(vehicle.queue) + 1):
                chain = list(vehicle.queue)
                chain.insert(pos, task_id)
                completions, empty_distance, finish = self._forecast_chain(
                    sim, vehicle, chain, now_s
                )
                flow = completions[pos] - sim.release_s[task_id]
                priority_weight = 1.0 + max(0, task.priority - 50) / 80.0
                score = sum(completions) - base_sum
                score += 0.12 * flow / priority_weight
                score += 1e-5 * (empty_distance - base_empty)
                # 空闲 OHT 可能正在巡航。若新任务的 Source 位于其当前方向上、
                # 且距离短于离散制动所需距离，立即改道会造成不可实现的急停。
                # 将这类空间接纳不可行的首任务候选施加大罚值，由其他车辆接单。
                if vehicle.current_task is None and not vehicle.queue and pos == 0:
                    empty_to_source, _ = self._travel(
                        (vehicle.link_id, vehicle.offset), task.source
                    )
                    braking_acceptance = (
                        vehicle.speed * core.DT
                        + vehicle.speed * vehicle.speed / (2.0 * core.DEC)
                        + core.STOP_LINE_BUFFER
                    )
                    if empty_to_source < braking_acceptance - core.EPS:
                        score += 1_000_000.0
                    if not self._route_admissible(sim, vehicle, task):
                        score += 1_000_000.0
                if self.q3_mode == "microbatch":
                    score += 1.5 * len(chain) + 0.5 * port_pressure
                    score += 0.01 * max(0.0, finish - now_s)
                else:
                    # 文档主模型：队列平方、热点Port、区域边际压力、时隙冲突和迟到风险。
                    score += 1.0 * len(chain) * len(chain)
                    score += 0.6 * port_pressure
                    score += 0.35 * marginal_pressure
                    score += 0.012 * max(0.0, finish - now_s)
                    if self.q3_mode == "full":
                        score += 0.8 * sim.estimate_slot_conflict(task)
                    if vehicle.state == "TO_DROPOFF":
                        score += 1.0
                candidate = (score, vehicle_id, pos)
                if best is None or candidate < best:
                    best = candidate
        assert best is not None
        return best[1], best[2], best[0]

    def dispatch(
        self, sim: "HighDensitySimulator", ready_ids: Sequence[str], now_s: float
    ) -> List[Dict[str, object]]:
        if self.q3_mode == "direct":
            return super().dispatch(sim, ready_ids, now_s)
        ordered = sorted(
            ready_ids,
            key=lambda x: (
                -self.data.tasks[x].priority,
                -(now_s - sim.release_s[x]),
                sim.release_s[x], x,
            ),
        )
        assignments: List[Dict[str, object]] = []
        for task_id in ordered:
            vehicle_id, pos, score = self._best_pressure_insertion(sim, task_id, now_s)
            sim.assign_task(task_id, vehicle_id, pos, now_s)
            assignments.append({
                "task_id": task_id,
                "vehicle_id": vehicle_id,
                "queue_position": pos,
                "bid": score,
            })
        return assignments


class HighDensitySimulator(online.OnlineTrafficSimulator):
    QUESTION_NO = 3

    def __init__(
        self,
        data: core.DataBundle,
        graph: core.GraphEngine,
        q3_mode: str,
        capture_trajectory: bool,
    ):
        self.q3_mode = q3_mode
        self.batch_interval_s = {
            "direct": 0.0,
            "microbatch": 5.0,
            "pressure": 3.0,
            "full": FULL_BATCH_S,
        }[q3_mode]
        self.last_batch_s = 0.0
        self.slot_calendar: Dict[str, List[Dict[str, Any]]] = defaultdict(list)
        self.slot_created_count = 0
        self.slot_entry_count = 0
        self.slot_shift_count = 0
        self.deadlock_recovery_log: List[Dict[str, Any]] = []
        self._last_slot_order_s = -math.inf
        self._last_deadlock_recovery_s = -math.inf
        self._last_cycle_signature: Tuple[str, ...] = ()
        self._cycle_persistence = 0
        online_mode = "document" if q3_mode == "full" else q3_mode
        super().__init__(data, graph, online_mode, capture_trajectory)

    def _initialize(self) -> None:
        self.scheduler = HighDensityScheduler(self.data, self.graph, self.q3_mode)
        super()._initialize()

    def link_region(self, link_id: int) -> str:
        node = self.data.links[link_id].to_node
        return f"R{(node - 1) // 10 + 1:02d}"

    def port_region(self, port_id: str) -> str:
        return self.link_region(self.data.ports[port_id].link_id)

    def region_pressure_snapshot(self) -> Dict[str, float]:
        regions = {self.link_region(link_id) for link_id in self.data.links}
        task_count: Counter[str] = Counter()
        vehicle_count: Counter[str] = Counter()
        idle_count: Counter[str] = Counter()
        occupied_length: Counter[str] = Counter()
        capacity_length: Counter[str] = Counter()
        resource_wait: Counter[str] = Counter()
        for link in self.data.links.values():
            capacity_length[self.link_region(link.link_id)] += link.length
        for vehicle in self.vehicles.values():
            region = self.link_region(vehicle.link_id)
            vehicle_count[region] += 1
            occupied_length[region] += core.REF_GAP
            if vehicle.current_task is None:
                idle_count[region] += 1
            ids = ([vehicle.current_task] if vehicle.current_task else []) + list(vehicle.queue)
            for task_id in ids:
                task = self.data.tasks[task_id]
                task_count[self.port_region(task.source)] += 0.5
                task_count[self.port_region(task.destination)] += 0.5
            if vehicle.pause_active and vehicle.pause_reason in {"汇流避让", "弯轨等待", "Port等待"}:
                resource_wait[region] += 1
        for task_id in self._ready_unassigned():
            task = self.data.tasks[task_id]
            task_count[self.port_region(task.source)] += 0.5
            task_count[self.port_region(task.destination)] += 0.5
        return {
            region: (
                task_count[region]
                + 0.55 * vehicle_count[region]
                + 8.0 * occupied_length[region] / max(core.REF_GAP, capacity_length[region])
                + 1.2 * resource_wait[region]
                - 0.35 * idle_count[region]
            )
            for region in regions
        }

    @staticmethod
    def _slot_key(kind: str, resource: object) -> str:
        return f"{kind}:{resource}"

    def _slot_duration(self, kind: str) -> float:
        if kind == "PORT":
            return 8.0
        return max(1.5, (core.REF_GAP + core.VEHICLE_LENGTH) / core.CONTROL_VMAX)

    def _allocate_slot(
        self,
        vehicle: core.Vehicle,
        kind: str,
        resource: object,
        eta_s: float,
    ) -> Dict[str, Any]:
        key = self._slot_key(kind, resource)
        duration = self._slot_duration(kind)
        start = max(self.time_s, eta_s)
        active = sorted(
            (x for x in self.slot_calendar[key] if x["status"] in {"reserved", "active"}),
            key=lambda x: (x["start_s"], x["vehicle_id"]),
        )
        for other in active:
            if start + duration + SLOT_GAP_S <= other["start_s"]:
                break
            if start < other["end_s"] + SLOT_GAP_S:
                start = other["end_s"] + SLOT_GAP_S
        slot = {
            "resource": key,
            "kind": kind,
            "resource_id": str(resource),
            "vehicle_id": vehicle.vehicle_id,
            "task_id": vehicle.current_task or "",
            "priority": self._task_priority(vehicle),
            "created_s": round(self.time_s, 3),
            "start_s": round(start, 3),
            "end_s": round(start + duration, 3),
            "actual_entry_s": None,
            "actual_exit_s": None,
            "status": "reserved",
        }
        self.slot_calendar[key].append(slot)
        self.slot_created_count += 1
        return slot

    def _active_slot(
        self, vehicle_id: str, kind: str, resource: object
    ) -> Optional[Dict[str, Any]]:
        key = self._slot_key(kind, resource)
        candidates = [
            x for x in self.slot_calendar.get(key, [])
            if x["vehicle_id"] == vehicle_id and x["status"] in {"reserved", "active"}
        ]
        return min(candidates, key=lambda x: x["start_s"]) if candidates else None

    def _plan_route_slots(self, vehicle: core.Vehicle) -> None:
        if self.q3_mode != "full" or vehicle.current_task is None:
            return
        for slots in self.slot_calendar.values():
            for slot in slots:
                if slot["vehicle_id"] == vehicle.vehicle_id and slot["status"] == "reserved":
                    slot["status"] = "cancelled"
        speed = max(vehicle.speed, 0.65 * core.CONTROL_VMAX)
        for kind, resource, distance in self._upcoming_controls(vehicle, horizon=20000.0):
            self._allocate_slot(vehicle, kind, resource, self.time_s + distance / speed)
        self._allocate_slot(
            vehicle, "PORT", vehicle.target_port,
            self.time_s + self._remaining_distance(vehicle) / speed,
        )

    def estimate_slot_conflict(self, task: core.Task) -> float:
        if self.q3_mode != "full":
            return 0.0
        keys = {self._slot_key("PORT", task.source), self._slot_key("PORT", task.destination)}
        return float(sum(
            1 for key in keys for slot in self.slot_calendar.get(key, [])
            if slot["status"] in {"reserved", "active"}
        ))

    def _route_to_port(self, vehicle: core.Vehicle, port_id: str, kind: str) -> None:
        super()._route_to_port(vehicle, port_id, kind)
        self._plan_route_slots(vehicle)

    def _arbitrate_resources(self) -> None:
        old_curve = dict(self.curve_owner)
        old_merge = dict(self.merge_owner)
        super()._arbitrate_resources()
        if self.q3_mode != "full":
            return
        for group, vehicle_id in old_curve.items():
            if self.curve_owner.get(group) != vehicle_id:
                slot = self._active_slot(vehicle_id, "CURVE", group)
                if slot is not None and slot["status"] == "active":
                    slot["status"] = "completed"
                    slot["actual_exit_s"] = round(self.time_s, 3)
        for node, vehicle_id in old_merge.items():
            if self.merge_owner.get(node) != vehicle_id:
                slot = self._active_slot(vehicle_id, "MERGE", node)
                if slot is not None and slot["status"] == "active":
                    slot["status"] = "completed"
                    slot["actual_exit_s"] = round(self.time_s, 3)

        if self.time_s - self._last_slot_order_s < 1.0 - core.EPS:
            return
        self._last_slot_order_s = self.time_s

        # 每1秒滚动补充有限时域时隙；实际放行仍由物理队首和资源容量仲裁，
        # 避免预测时隙覆盖已经接近停止线的车辆权限。
        for vehicle in self.vehicles.values():
            if vehicle.state not in self.MOVING_STATES or vehicle.current_task is None:
                continue
            for kind, resource, distance in self._upcoming_controls(vehicle, horizon=6000.0):
                slot = self._active_slot(vehicle.vehicle_id, kind, resource)
                if slot is None:
                    speed = max(vehicle.speed, 0.65 * core.CONTROL_VMAX)
                    self._allocate_slot(
                        vehicle, kind, resource, self.time_s + distance / speed
                    )

    def _mark_slot_entry(
        self, vehicle: core.Vehicle, kind: str, resource: object, at_s: float
    ) -> None:
        slot = self._active_slot(vehicle.vehicle_id, kind, resource)
        if slot is None:
            slot = self._allocate_slot(vehicle, kind, resource, at_s)
            self.slot_shift_count += 1
        if at_s > slot["end_s"] + core.EPS:
            duration = slot["end_s"] - slot["start_s"]
            slot["start_s"] = round(at_s, 3)
            slot["end_s"] = round(at_s + duration, 3)
            self.slot_shift_count += 1
        slot["actual_entry_s"] = round(at_s, 3)
        slot["status"] = "active"
        self.slot_entry_count += 1

    def _cross_boundary(self, vehicle: core.Vehicle, old_link_id: int, next_link_id: int) -> None:
        super()._cross_boundary(vehicle, old_link_id, next_link_id)
        if self.q3_mode != "full" or vehicle.current_task is None:
            return
        old_link = self.data.links[old_link_id]
        next_link = self.data.links[next_link_id]
        if old_link.to_node in self.data.merge_nodes:
            self._mark_slot_entry(vehicle, "MERGE", old_link.to_node, self.time_s)
        if next_link.curve_group and next_link.curve_group != old_link.curve_group:
            self._mark_slot_entry(vehicle, "CURVE", next_link.curve_group, self.time_s)

    def _handle_arrivals(self, at_s: float) -> None:
        before = {v.vehicle_id: v.state for v in self.vehicles.values()}
        super()._handle_arrivals(at_s)
        if self.q3_mode != "full":
            return
        for vehicle in self.vehicles.values():
            if before.get(vehicle.vehicle_id) in {"TO_PICKUP", "TO_DROPOFF"} and vehicle.state in {"PICKING", "DROPPING"}:
                self._mark_slot_entry(vehicle, "PORT", vehicle.target_port, at_s)

    def _process_services(self, end_s: float) -> None:
        before = {
            v.vehicle_id: (v.state, v.target_port)
            for v in self.vehicles.values() if v.state in {"PICKING", "DROPPING"}
        }
        super()._process_services(end_s)
        if self.q3_mode != "full":
            return
        for vehicle_id, (old_state, port_id) in before.items():
            if self.vehicles[vehicle_id].state == old_state:
                continue
            slot = self._active_slot(vehicle_id, "PORT", port_id)
            if slot is not None and slot["status"] == "active":
                slot["status"] = "completed"
                slot["actual_exit_s"] = round(end_s, 3)

    def _update_wait_graph(self) -> None:
        if self.q3_mode != "full":
            super()._update_wait_graph()
            return
        if self.time_s - self._last_wait_graph_check_s < 1.0 - core.EPS:
            return
        edges: Dict[str, set[str]] = defaultdict(set)
        waiting_reasons = {"安全跟驰", "汇流避让", "弯轨等待", "Port等待"}
        for vehicle in self.vehicles.values():
            if vehicle.speed > core.EPS or vehicle.last_reason not in waiting_reasons:
                continue
            blocked, _, owner = self._boundary_block(vehicle)
            if blocked and owner and owner != vehicle.vehicle_id:
                edges[vehicle.vehicle_id].add(owner)
            front = self._front_ahead(vehicle)
            if front is not None and front[0].speed <= core.EPS and front[1] <= 2.0 * core.REF_GAP:
                edges[vehicle.vehicle_id].add(front[0].vehicle_id)
        self.wait_graph_max_edges = max(self.wait_graph_max_edges, sum(len(v) for v in edges.values()))
        cycle = self._wait_cycle(edges)
        self._last_wait_graph_check_s = self.time_s
        if not cycle:
            self._last_cycle_signature = ()
            self._cycle_persistence = 0
            return
        signature = tuple(sorted(set(cycle[:-1])))
        if signature == self._last_cycle_signature:
            self._cycle_persistence += 1
        else:
            self._last_cycle_signature = signature
            self._cycle_persistence = 1
        if self._cycle_persistence < 5:
            return
        self.wait_graph_cycle_count += 1
        self.pending_deadlock_trigger = True
        if self.time_s - self._last_deadlock_recovery_s < 30.0 - core.EPS:
            return
        candidates: List[Tuple[int, str, str, object]] = []
        participants = set(signature)
        for group, vehicle_id in self.curve_owner.items():
            vehicle = self.vehicles[vehicle_id]
            inside = self.data.links[vehicle.link_id].curve_group == group
            waited = self.time_s - self.resource_wait_since.get(vehicle_id, self.time_s)
            if vehicle_id in participants and not inside and group not in self.curve_release and vehicle.state != "TO_DROPOFF" and waited >= 10.0:
                candidates.append((self._task_priority(vehicle), vehicle_id, "CURVE", group))
        for node, vehicle_id in self.merge_owner.items():
            vehicle = self.vehicles[vehicle_id]
            waited = self.time_s - self.resource_wait_since.get(vehicle_id, self.time_s)
            if vehicle_id in participants and node not in self.merge_release and vehicle.state != "TO_DROPOFF" and waited >= 10.0:
                candidates.append((self._task_priority(vehicle), vehicle_id, "MERGE", node))
        if not candidates:
            return
        priority, vehicle_id, kind, resource = min(candidates)
        if kind == "CURVE":
            self.curve_owner.pop(str(resource), None)
        else:
            self.merge_owner.pop(int(resource), None)
        cancelled = []
        for slots in self.slot_calendar.values():
            for slot in slots:
                if slot["vehicle_id"] == vehicle_id and slot["status"] == "reserved":
                    slot["status"] = "cancelled"
                    cancelled.append(slot["resource"])
        self._plan_route_slots(self.vehicles[vehicle_id])
        self.deadlock_recovery_log.append({
            "time_s": round(self.time_s, 3),
            "vehicle_id": vehicle_id,
            "priority": priority,
            "action": "撤销最低优先级未生效预约并重新排时隙",
            "cancelled_resources": cancelled,
        })
        self._last_deadlock_recovery_s = self.time_s
        self._cycle_persistence = 0

    def _record_decision(
        self,
        now_s: float,
        trigger: str,
        new_ids: Sequence[str],
        considered: Sequence[str],
        assignments: Sequence[Dict[str, object]],
    ) -> None:
        max_visible = max((self.release_s[x] for x in self.released_ids), default=-math.inf)
        if max_visible > now_s + core.EPS:
            raise RuntimeError("在线信息泄漏")
        self.decision_log.append({
            "decision_time_s": round(now_s, 3),
            "trigger": trigger,
            "new_task_ids": sorted(new_ids),
            "visible_task_count": len(self.released_ids),
            "backlog_count": len(self._ready_unassigned()),
            "considered_task_ids": sorted(considered),
            "max_visible_release_s": round(max_visible, 3),
            "assignments": list(assignments),
            "region_pressure": {
                key: round(value, 6)
                for key, value in sorted(self.region_pressure_snapshot().items())
            },
            "slot_reservation_count": self.slot_created_count,
            "future_leak": False,
        })

    def _release_and_dispatch(self, now_s: float, trigger: str) -> None:
        new_ids = self._release_tasks(now_s)
        ready = self._ready_unassigned()
        if not ready:
            return
        if self.q3_mode == "direct":
            heartbeat = now_s - self.last_dispatch_s >= online.HEARTBEAT_S - core.EPS
            if not new_ids and not heartbeat and trigger != "completion":
                return
            assignments = self.scheduler.dispatch(self, ready, now_s)
            self._record_decision(
                now_s, "release" if new_ids else trigger, new_ids, ready, assignments
            )
            self.last_dispatch_s = now_s
            return

        high = [x for x in ready if self.data.tasks[x].priority >= HIGH_PRIORITY]
        batch_due = now_s - self.last_batch_s >= self.batch_interval_s - core.EPS
        count_due = self.q3_mode == "full" and len(ready) >= FULL_BATCH_COUNT
        oldest_age = max(now_s - self.release_s[x] for x in ready)
        idle_available = any(v.current_task is None for v in self.vehicles.values())
        early_idle = idle_available and oldest_age >= self.batch_interval_s / 2.0
        if batch_due or count_due or early_idle:
            considered = ready
            event = "microbatch" if batch_due else ("batch_count" if count_due else "idle_capacity")
        elif high:
            considered = high
            event = "high_priority"
        else:
            return
        assignments = self.scheduler.dispatch(self, considered, now_s)
        self._record_decision(now_s, event, new_ids, considered, assignments)
        self.last_dispatch_s = now_s
        if batch_due:
            self.last_batch_s = now_s

    def run(self, max_time_s: float = MAX_TIME_S) -> Dict[str, object]:
        summary = super().run(max_time_s)
        summary["label"] = self.q3_mode
        summary["batch_interval_s"] = self.batch_interval_s
        summary["max_ready_backlog"] = max(
            (int(x.get("backlog_count", 0)) for x in self.decision_log), default=0
        )
        pressures = [
            max((float(v) for v in x.get("region_pressure", {}).values()), default=0.0)
            for x in self.decision_log
        ]
        summary["max_region_pressure"] = max(pressures, default=0.0)
        summary["slot_reservations_created"] = self.slot_created_count
        summary["slot_entries"] = sum(
            slot["actual_entry_s"] is not None
            for slots in self.slot_calendar.values() for slot in slots
        )
        summary["slot_reservations_shifted"] = self.slot_shift_count
        summary["reservation_event_count"] = len(self.reservation_log)
        summary["wait_graph_cycle_count"] = self.wait_graph_cycle_count
        summary["wait_graph_max_edges"] = self.wait_graph_max_edges
        summary["deadlock_recovery_count"] = len(self.deadlock_recovery_log)
        return summary


def absolute_time(base: datetime, relative_s: Optional[float]) -> Optional[datetime]:
    if relative_s is None:
        return None
    return base + timedelta(milliseconds=round(relative_s * 1000.0))


def task_rows(sim: HighDensitySimulator) -> List[List[object]]:
    rows: List[List[object]] = []
    for task in sorted(sim.data.tasks.values(), key=lambda t: (t.install_time, t.command_id)):
        record = sim.records[task.command_id]
        release = sim.release_s[task.command_id]
        if record.completed_s is None:
            raise RuntimeError(f"任务未完成: {task.command_id}")
        rows.append([
            3, task.command_id, task.carrier_id, task.priority, task.install_time,
            task.source, task.destination, record.vehicle_id,
            absolute_time(sim.base_time, record.assigned_s),
            int(round((record.assigned_s - release) * 1000.0)),
            absolute_time(sim.base_time, record.source_arrived_s),
            absolute_time(sim.base_time, record.acquire_start_s),
            absolute_time(sim.base_time, record.acquire_end_s),
            absolute_time(sim.base_time, record.departed_s),
            absolute_time(sim.base_time, record.destination_arrived_s),
            absolute_time(sim.base_time, record.deposit_start_s),
            absolute_time(sim.base_time, record.deposit_end_s),
            absolute_time(sim.base_time, record.completed_s),
            round(record.to_source_distance, 3), round(record.to_destination_distance, 3),
            round(record.completed_s - release, 3), record.move_to_source_path,
            record.move_to_destination_path, record.paused_count, record.pausing_ms,
        ])
    return rows


def metric_row(sim: HighDensitySimulator, runtime_s: float) -> List[object]:
    records = list(sim.records.values())
    return [
        3, 600,
        round(statistics.mean(r.completed_s - sim.release_s[r.task.command_id] for r in records), 6),
        round(statistics.mean(r.assigned_s - sim.release_s[r.task.command_id] for r in records), 6),
        round(statistics.mean(r.source_arrived_s - r.assigned_s for r in records), 6),
        round(statistics.mean(r.destination_arrived_s - r.departed_s for r in records), 6),
        round(statistics.mean(r.pausing_ms / 1000.0 for r in records), 6),
        round(max(r.completed_s for r in records), 6),
        round(statistics.mean(r.to_source_distance for r in records), 6),
        round(statistics.mean(r.to_source_distance + r.to_destination_distance for r in records), 6),
        round(runtime_s, 6),
    ]


def export_excel(
    root: Path,
    out_path: Path,
    tasks: List[List[object]],
    trajectory: List[List[object]],
    metrics: List[object],
) -> None:
    wb = openpyxl.load_workbook(root / "附件9_结果提交模板.xlsx")
    ws_task = wb["任务仿真结果"]
    ws_trace = wb["OHT逐步运行记录表"]
    ws_metric = wb["算法评价指标"]
    q3_rows = [row for row in range(2, ws_task.max_row + 1) if ws_task.cell(row, 1).value == 3]
    if len(q3_rows) != 600:
        raise ValueError(f"附件9问题3预置行 {len(q3_rows)} != 600")
    for row_idx, values in zip(q3_rows, tasks):
        for col_idx, value in enumerate(values, start=1):
            cell = ws_task.cell(row_idx, col_idx, value)
            if isinstance(value, datetime):
                cell.number_format = "yyyy-mm-dd hh:mm:ss.000"
    for row in trajectory:
        if int(row[0]) != 3:
            raise ValueError(f"问题3轨迹问题号错误: {row[0]}")
        ws_trace.append(row)
    for col_idx, value in enumerate(metrics, start=1):
        ws_metric.cell(4, col_idx, value)
    wb.save(out_path)
    check = openpyxl.load_workbook(out_path, data_only=False, read_only=True)
    if check["任务仿真结果"].max_row != 823:
        raise ValueError("任务模板行数被破坏")
    if check["OHT逐步运行记录表"].max_row != len(trajectory) + 1:
        raise ValueError("轨迹写入行数不一致")
    if check["算法评价指标"].cell(4, 2).value != 600:
        raise ValueError("问题3指标未正确写入")
    if any(int(row[0].value) != 3 for row in check["OHT逐步运行记录表"].iter_rows(min_row=2, max_col=1)):
        raise ValueError("问题3 Excel轨迹问题号错误")


def audit_result(
    data: core.DataBundle, sim: HighDensitySimulator, summary: Dict[str, object]
) -> Dict[str, object]:
    errors: List[str] = []
    for task_id in data.tasks:
        if task_id not in sim.records:
            errors.append(f"{task_id}: 未分配")
            continue
        r = sim.records[task_id]
        release = sim.release_s[task_id]
        values = [
            release, r.assigned_s, r.source_arrived_s, r.acquire_start_s,
            r.acquire_end_s, r.departed_s, r.destination_arrived_s,
            r.deposit_start_s, r.deposit_end_s, r.completed_s,
        ]
        if any(x is None for x in values):
            errors.append(f"{task_id}: 时间字段不完整")
            continue
        if any(a > b + core.EPS for a, b in zip(values, values[1:])):
            errors.append(f"{task_id}: 时间非单调")
        if abs((r.acquire_end_s - r.acquire_start_s) - core.PICK_TIME) > 1e-6:
            errors.append(f"{task_id}: 取货时间错误")
        if abs((r.deposit_end_s - r.deposit_start_s) - core.DROP_TIME) > 1e-6:
            errors.append(f"{task_id}: 放货时间错误")
        pred = sim.predecessor[task_id]
        if pred is not None:
            previous = sim.records.get(pred)
            if previous is None or previous.completed_s > r.assigned_s + core.EPS:
                errors.append(f"{task_id}: Carrier前序约束错误")
    online_errors = []
    for decision in sim.decision_log:
        if decision["max_visible_release_s"] > decision["decision_time_s"] + core.EPS:
            online_errors.append(f"{decision['decision_time_s']}: 读取未来任务")
        if decision.get("future_leak"):
            online_errors.append(f"{decision['decision_time_s']}: future_leak")
    step_counts = Counter(int(row[1]) for row in sim.trajectory)
    bad_steps = [s for s, count in step_counts.items() if count != 20]
    result = {
        "data_audit": data.audit,
        "simulation_summary": summary,
        "task_validation_errors": errors,
        "online_validation_errors": online_errors,
        "bad_trajectory_steps": bad_steps,
        "traffic_violations": sim.violations,
        "carrier_precedence_edges": sum(1 for x in sim.predecessor.values() if x is not None),
        "decision_count": len(sim.decision_log),
        "slot_audit": {
            "created": sim.slot_created_count,
            "entries": sim.slot_entry_count,
            "shifted": sim.slot_shift_count,
            "actual_overlap_errors": [],
        },
        "wait_graph_audit": {
            "cycle_count": sim.wait_graph_cycle_count,
            "max_edges": sim.wait_graph_max_edges,
            "recovery_count": len(sim.deadlock_recovery_log),
            "recovery_events": sim.deadlock_recovery_log,
        },
    }
    for resource, slots in sim.slot_calendar.items():
        used = sorted(
            (x for x in slots if x["actual_entry_s"] is not None),
            key=lambda x: x["actual_entry_s"],
        )
        for left, right in zip(used, used[1:]):
            if (
                left["actual_exit_s"] is not None
                and right["actual_entry_s"] < left["actual_exit_s"] - core.EPS
            ):
                result["slot_audit"]["actual_overlap_errors"].append(
                    f"{resource}: {left['vehicle_id']} 与 {right['vehicle_id']}"
                )
    result["passed"] = (
        summary["completed_tasks"] == 600
        and summary["hard_violation_count"] == 0
        and summary["future_leak_count"] == 0
        and not errors and not online_errors and not bad_steps
        and not result["slot_audit"]["actual_overlap_errors"]
    )
    return result


def write_report(
    path: Path,
    summaries: Dict[str, Dict[str, object]],
    selected: str,
    metrics: Sequence[object],
) -> None:
    labels = {
        "direct": "直接复用问题2滚动插入",
        "microbatch": "5秒普通微批",
        "pressure": "3秒压力感知微批（无显式时隙）",
        "full": "文档主模型：压力微批+时隙预约+防死锁",
    }
    lines = [
        "# 问题3：600项高密动态任务结果\n\n",
        "严格按全过程计划采用压力感知微批、区域压力平衡、MERGE/CURVE/Port有限时域时隙、空间接纳和等待图防死锁。直接复用问题2只作为基准，调度器不读取未来任务。\n\n",
        "| 算法 | 平均任务执行时间/s | Makespan/s | 硬违规 |\n",
        "|---|---:|---:|---:|\n",
    ]
    for mode in ["direct", "microbatch", "pressure", "full"]:
        s = summaries[mode]
        lines.append(
            f"| {labels[mode]} | {s['avg_transfer_time_s']:.6f} | {s['makespan_s']:.3f} | {s['hard_violation_count']} |\n"
        )
    lines.extend([
        f"\n最终选择：**{labels[selected]}**。\n\n",
        "## 正式指标\n\n",
        f"- AvgTransferTime：{metrics[2]} s\n",
        f"- AvgAssignTime：{metrics[3]} s\n",
        f"- AvgPickupResponseTime：{metrics[4]} s\n",
        f"- AvgTransportTime：{metrics[5]} s\n",
        f"- AvgPausingTime：{metrics[6]} s\n",
        f"- Makespan：{metrics[7]} s\n",
        f"- AvgEmptyDistance：{metrics[8]} mm\n",
        f"- AvgTotalDistance：{metrics[9]} mm\n\n",
        "```powershell\npython .\\q3_solver.py\npython .\\verify_q3.py\n```\n",
    ])
    path.write_text("".join(lines), encoding="utf-8")


def solve(root: Path) -> Dict[str, object]:
    out = root / "outputs" / "q3"
    out.mkdir(parents=True, exist_ok=True)
    data = load_q3_data(root)
    graph = core.GraphEngine(data)
    summaries: Dict[str, Dict[str, object]] = {}
    for mode in ["direct", "microbatch", "pressure", "full"]:
        print(f"运行问题3 {mode} 无轨迹对比仿真……", flush=True)
        sim = HighDensitySimulator(data, graph, mode, capture_trajectory=False)
        summaries[mode] = sim.run()
        summaries[mode]["hard_gate_passed"] = (
            summaries[mode]["completed_tasks"] == 600
            and summaries[mode]["hard_violation_count"] == 0
            and summaries[mode]["future_leak_count"] == 0
            and summaries[mode]["min_continuous_clearance_mm"] is not None
            and float(summaries[mode]["min_continuous_clearance_mm"]) >= core.CLEAR_GAP - 1e-4
        )
        print(json.dumps(summaries[mode], ensure_ascii=False, indent=2), flush=True)
    feasible = [mode for mode, summary in summaries.items() if summary["hard_gate_passed"] is True]
    if not feasible:
        (out / "问题3_失败诊断.json").write_text(
            json.dumps(summaries, ensure_ascii=False, indent=2), encoding="utf-8"
        )
        raise RuntimeError("问题3候选算法均未得到零违规完整解")
    selected = min(
        feasible,
        key=lambda mode: (
            summaries[mode]["avg_transfer_time_s"],
            summaries[mode]["makespan_s"],
            mode,
        ),
    )
    print(f"全日志复跑最终方案：{selected}……", flush=True)
    sim = HighDensitySimulator(data, graph, selected, capture_trajectory=True)
    final_summary = sim.run()
    summaries[selected] = final_summary
    rows = task_rows(sim)
    metrics = metric_row(sim, final_summary["simulation_runtime_s"])
    audit = audit_result(data, sim, final_summary)
    if not audit["passed"]:
        (out / "问题3_失败诊断.json").write_text(
            json.dumps(audit, ensure_ascii=False, indent=2), encoding="utf-8"
        )
        raise RuntimeError("问题3最终方案未通过完整审计")
    stale = out / "问题3_失败诊断.json"
    if stale.exists():
        stale.unlink()

    core.write_csv(out / "问题3_任务结果.csv", core.TASK_HEADERS, rows)
    core.write_csv(out / "问题3_OHT逐步运行记录.csv", core.TRAJECTORY_HEADERS, sim.trajectory)
    core.write_csv(out / "问题3_算法评价指标.csv", core.METRIC_HEADERS, [metrics])
    (out / "问题3_在线决策日志.json").write_text(
        json.dumps(sim.decision_log, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    (out / "问题3_资源时隙日志.json").write_text(
        json.dumps(
            [slot for key in sorted(sim.slot_calendar) for slot in sim.slot_calendar[key]],
            ensure_ascii=False, indent=2,
        ), encoding="utf-8"
    )
    (out / "问题3_资源预约日志.json").write_text(
        json.dumps(sim.reservation_log, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    (out / "问题3_防死锁日志.json").write_text(
        json.dumps({
            "wait_graph_cycle_count": sim.wait_graph_cycle_count,
            "wait_graph_max_edges": sim.wait_graph_max_edges,
            "recovery_events": sim.deadlock_recovery_log,
        }, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    (out / "问题3_约束审计.json").write_text(
        json.dumps(audit, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    (out / "问题3_方案.json").write_text(
        json.dumps({
            "selected_mode": selected,
            "summaries": summaries,
            "configuration": {
                "dt_s": core.DT,
                "control_vmax_mm_s": core.CONTROL_VMAX,
                "microbatch_s": {"direct": 0, "microbatch": 5, "pressure": 3, "full": FULL_BATCH_S},
                "microbatch_count_trigger": FULL_BATCH_COUNT,
                "high_priority_immediate_threshold": HIGH_PRIORITY,
                "region_partition": "按Link下游节点编号每10个节点一组",
                "explicit_slot_resources": ["MERGE", "CURVE", "PORT"],
                "slot_gap_s": SLOT_GAP_S,
                "deadlock_recovery": "撤销最低优先级未生效预约并重新排时隙",
                "clear_gap_mm": core.CLEAR_GAP,
                "reference_gap_mm": core.REF_GAP,
                "future_task_fields_visible_to_scheduler": False,
                "no_reassignment_after_assignment": True,
                "p17_override": [70, 507, 303.0],
            },
        }, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    export_excel(root, out / "问题3_结果.xlsx", rows, sim.trajectory, metrics)
    write_report(out / "问题3_结果说明.md", summaries, selected, metrics)
    return {
        "selected": selected,
        "metrics": dict(zip(core.METRIC_HEADERS, metrics)),
        "audit_passed": True,
        "output_dir": str(out),
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="求解OHT赛题问题3")
    parser.add_argument("--root", type=Path, default=Path(__file__).resolve().parents[1])
    args = parser.parse_args()
    print(json.dumps(solve(args.root.resolve()), ensure_ascii=False, indent=2), flush=True)


if __name__ == "__main__":
    main()
