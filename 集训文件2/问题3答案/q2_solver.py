from __future__ import annotations

import argparse
import csv
import json
import math
import statistics
import time
from collections import Counter, defaultdict
from dataclasses import asdict
from datetime import datetime, timedelta
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Sequence, Tuple

import openpyxl

import q1_solver as core


HEARTBEAT_S = 2.0
MAX_TIME_S = 9000.0
DOCUMENT_HEARTBEAT_S = 3.0
MIN_REPLAN_INTERVAL_S = 1.0
COMMIT_LOCK_S = 8.0
REASSIGN_THRESHOLD_S = 300.0
REASSIGN_COOLDOWN_S = 10.0
RESOURCE_DELAY_TRIGGER_S = 5.0


def load_q2_data(root: Path) -> core.DataBundle:
    data = core.DataBundle(root).load()
    ws = openpyxl.load_workbook(
        root / "附件6_动态任务数据.xlsx", data_only=True, read_only=True
    )["Task"]
    tasks: Dict[str, core.Task] = {}
    for command_id, install_time, source, destination, priority, carrier_id in ws.iter_rows(
        min_row=2, values_only=True
    ):
        task = core.Task(
            str(command_id), install_time, str(source), str(destination),
            int(priority), str(carrier_id),
        )
        tasks[task.command_id] = task
    if len(tasks) != 190:
        raise ValueError(f"问题2任务数量 {len(tasks)} != 190")
    for task in tasks.values():
        if task.source not in data.ports or task.destination not in data.ports:
            raise ValueError(f"{task.command_id} Port 外键无效")
    data.tasks = tasks
    carrier_count = Counter(t.carrier_id for t in tasks.values())
    base = min(t.install_time for t in tasks.values())
    data.audit = {
        **data.audit,
        "counts": {**data.audit["counts"], "tasks": 190},
        "scenario": "问题2",
        "release_span_s": (max(t.install_time for t in tasks.values()) - base).total_seconds(),
        "different_carriers": len(carrier_count),
        "repeated_carriers": sum(v > 1 for v in carrier_count.values()),
        "max_carrier_chain": max(carrier_count.values()),
        "errors": [],
        "passed": True,
    }
    return data


class OnlineScheduler:
    """只接收仿真器传入的已释放、已解锁任务 ID。"""

    def __init__(self, data: core.DataBundle, graph: core.GraphEngine, mode: str):
        self.data = data
        self.graph = graph
        self.mode = mode
        self._travel_cache: Dict[Tuple[str, str], Tuple[float, float]] = {}

    def _travel(self, origin: Tuple[int, float] | str, destination: str) -> Tuple[float, float]:
        if isinstance(origin, tuple):
            key = (f"L{origin[0]}@{origin[1]:.1f}", destination)
            link_id, offset = origin
        else:
            key = (origin, destination)
            p = self.data.ports[origin]
            link_id, offset = p.link_id, p.offset
        if key not in self._travel_cache:
            _, distance, _, _ = self.graph.position_path(
                link_id, offset, self.data.ports[destination]
            )
            self._travel_cache[key] = (distance, distance / core.CONTROL_VMAX)
        return self._travel_cache[key]

    def _base_after_current(
        self, sim: "OnlineTrafficSimulator", vehicle: core.Vehicle, now_s: float
    ) -> Tuple[float, Tuple[int, float] | str]:
        if vehicle.current_task is None:
            return now_s, (vehicle.link_id, vehicle.offset)
        task = self.data.tasks[vehicle.current_task]
        if vehicle.state == "TO_PICKUP":
            remaining = sim._remaining_distance(vehicle) / core.CONTROL_VMAX
            _, loaded = self._travel(task.source, task.destination)
            finish = now_s + remaining + core.PICK_TIME + loaded + core.DROP_TIME
        elif vehicle.state == "PICKING":
            _, loaded = self._travel(task.source, task.destination)
            finish = now_s + vehicle.service_remaining + loaded + core.DROP_TIME
        elif vehicle.state == "TO_DROPOFF":
            finish = now_s + sim._remaining_distance(vehicle) / core.CONTROL_VMAX + core.DROP_TIME
        else:
            finish = now_s + vehicle.service_remaining
        return finish, task.destination

    def _forecast_chain(
        self,
        sim: "OnlineTrafficSimulator",
        vehicle: core.Vehicle,
        chain: Sequence[str],
        now_s: float,
    ) -> Tuple[List[float], float, float]:
        current_time, origin = self._base_after_current(sim, vehicle, now_s)
        completions: List[float] = []
        empty_distance = 0.0
        for task_id in chain:
            task = self.data.tasks[task_id]
            de, te = self._travel(origin, task.source)
            _, tl = self._travel(task.source, task.destination)
            current_time += te + core.PICK_TIME + tl + core.DROP_TIME
            empty_distance += de
            completions.append(current_time)
            origin = task.destination
        return completions, empty_distance, current_time

    def _port_pressure(self, sim: "OnlineTrafficSimulator", task: core.Task) -> float:
        count = 0
        for vehicle in sim.vehicles.values():
            ids = ([vehicle.current_task] if vehicle.current_task else []) + list(vehicle.queue)
            for task_id in ids:
                planned = self.data.tasks[task_id]
                if planned.source in {task.source, task.destination}:
                    count += 1
                if planned.destination in {task.source, task.destination}:
                    count += 1
        return float(count)

    def _historical_congestion_risk(
        self, sim: "OnlineTrafficSimulator", task: core.Task
    ) -> float:
        """只用当前和历史状态估计拥堵，不读取未释放任务。"""
        pressure = self._port_pressure(sim, task)
        resource_waits = sum(
            1 for vehicle in sim.vehicles.values()
            if vehicle.pause_active and vehicle.pause_reason in {"安全跟驰", "汇流避让", "弯轨等待", "Port等待"}
        )
        active_resources = len(sim.port_owner) + len(sim.curve_owner) + len(sim.merge_owner)
        completed_pause_s = sum(record.pausing_ms for record in sim.records.values()) / 1000.0
        history_scale = completed_pause_s / max(1, len(sim.records))
        return pressure + 0.5 * resource_waits + 0.1 * active_resources + 0.02 * history_scale

    def _route_admissible(
        self, sim: "OnlineTrafficSimulator", vehicle: core.Vehicle, task: core.Task
    ) -> bool:
        """空闲巡航车改道前检查前车与短 Link 后控制边界的可制动接纳距离。"""
        route, _, _, _ = self.graph.position_path(
            vehicle.link_id, vehicle.offset, self.data.ports[task.source]
        )
        required_stop = vehicle.speed * core.DT + vehicle.speed * vehicle.speed / (2.0 * core.DEC)
        cumulative = 0.0
        for idx, segment in enumerate(route):
            start = vehicle.offset if idx == 0 else segment.start
            for other in sim.vehicles.values():
                if other.vehicle_id == vehicle.vehicle_id or other.link_id != segment.link_id:
                    continue
                if other.offset < start - core.EPS:
                    continue
                distance = cumulative + other.offset - start
                if distance < core.REF_GAP + core.CONTINUOUS_GAP_MARGIN + required_stop - core.EPS:
                    return False
            cumulative += segment.end - start
            if idx + 1 >= len(route):
                continue
            link = self.data.links[segment.link_id]
            nxt = self.data.links[route[idx + 1].link_id]
            if link.to_node in self.data.merge_nodes:
                owner = sim.merge_owner.get(link.to_node)
                if owner not in {None, vehicle.vehicle_id} and cumulative < core.REF_GAP + required_stop - core.EPS:
                    return False
            if nxt.curve_group and nxt.curve_group != link.curve_group:
                owner = sim.curve_owner.get(nxt.curve_group)
                if owner not in {None, vehicle.vehicle_id} and cumulative < core.STOP_LINE_BUFFER + required_stop - core.EPS:
                    return False
        return True

    def _idle_task_admissible(
        self, sim: "OnlineTrafficSimulator", vehicle: core.Vehicle, task: core.Task
    ) -> bool:
        """高速空闲车只有在本步即可物理制动到 Source 时才允许立即接单。"""
        empty_to_source, _ = self._travel((vehicle.link_id, vehicle.offset), task.source)
        braking_acceptance = (
            vehicle.speed * core.DT
            + vehicle.speed * vehicle.speed / (2.0 * core.DEC)
            + core.STOP_LINE_BUFFER
        )
        return (
            empty_to_source >= braking_acceptance - core.EPS
            and self._route_admissible(sim, vehicle, task)
        )

    def _best_document_insertion(
        self, sim: "OnlineTrafficSimulator", task_id: str, now_s: float
    ) -> Tuple[str, int, float, float]:
        task = self.data.tasks[task_id]
        congestion = self._historical_congestion_risk(sim, task)
        best: Optional[Tuple[float, str, int, float]] = None
        for vehicle_id in sorted(sim.vehicles):
            vehicle = sim.vehicles[vehicle_id]
            base_comp, base_empty, _ = self._forecast_chain(sim, vehicle, vehicle.queue, now_s)
            base_sum = sum(base_comp)
            for pos in range(len(vehicle.queue) + 1):
                chain = list(vehicle.queue)
                chain.insert(pos, task_id)
                completions, empty_distance, finish = self._forecast_chain(sim, vehicle, chain, now_s)
                predicted_completion = completions[pos]
                flow = predicted_completion - sim.release_s[task_id]
                delta_empty = empty_distance - base_empty
                predicted_wait = max(0.0, finish - now_s - sum(
                    self._travel(
                        self.data.tasks[x].source,
                        self.data.tasks[x].destination,
                    )[1] + core.PICK_TIME + core.DROP_TIME
                    for x in chain
                ))
                score = flow
                score += 0.00005 * delta_empty
                score += 0.04 * predicted_wait
                score += 0.8 * congestion
                score -= 0.04 * task.priority
                score -= 0.08 * max(0.0, now_s - sim.release_s[task_id])
                score += 0.02 * (sum(completions) - base_sum)
                if vehicle.current_task is None and not vehicle.queue and pos == 0:
                    if not self._idle_task_admissible(sim, vehicle, task):
                        score += 1_000_000.0
                candidate = (score, vehicle_id, pos, predicted_completion)
                if best is None or candidate < best:
                    best = candidate
        assert best is not None
        return best[1], best[2], best[0], best[3]

    def _document_auction(
        self, sim: "OnlineTrafficSimulator", ready_ids: Sequence[str], now_s: float
    ) -> List[Dict[str, object]]:
        """滚动窗口内重复求最小报价，形成小规模指派—排序拍卖。"""
        pending = set(ready_ids)
        assignments: List[Dict[str, object]] = []
        while pending:
            options = []
            for task_id in sorted(pending):
                vehicle_id, pos, score, completion = self._best_document_insertion(sim, task_id, now_s)
                task = self.data.tasks[task_id]
                options.append((score, -task.priority, sim.release_s[task_id], task_id, vehicle_id, pos, completion))
            score, _, _, task_id, vehicle_id, pos, completion = min(options)
            sim.assign_task(task_id, vehicle_id, pos, now_s)
            assignments.append({
                "task_id": task_id,
                "vehicle_id": vehicle_id,
                "queue_position": pos,
                "bid": score,
                "predicted_completion_s": completion,
            })
            pending.remove(task_id)
        return assignments

    def _best_baseline(
        self, sim: "OnlineTrafficSimulator", task_id: str, now_s: float
    ) -> Tuple[str, int, float]:
        task = self.data.tasks[task_id]
        best: Optional[Tuple[float, str, int]] = None
        for vehicle_id in sorted(sim.vehicles):
            vehicle = sim.vehicles[vehicle_id]
            completions, _, available = self._forecast_chain(sim, vehicle, vehicle.queue, now_s)
            origin: Tuple[int, float] | str
            if vehicle.queue:
                origin = self.data.tasks[vehicle.queue[-1]].destination
            elif vehicle.current_task:
                origin = self.data.tasks[vehicle.current_task].destination
            else:
                origin = (vehicle.link_id, vehicle.offset)
            _, empty_t = self._travel(origin, task.source)
            score = available + empty_t
            if (
                vehicle.current_task is None
                and not vehicle.queue
                and not self._idle_task_admissible(sim, vehicle, task)
            ):
                score += 1_000_000.0
            candidate = (score, vehicle_id, len(vehicle.queue))
            if best is None or candidate < best:
                best = candidate
        assert best is not None
        return best[1], best[2], best[0]

    def _best_insertion(
        self, sim: "OnlineTrafficSimulator", task_id: str, now_s: float
    ) -> Tuple[str, int, float]:
        task = self.data.tasks[task_id]
        best: Optional[Tuple[float, str, int]] = None
        for vehicle_id in sorted(sim.vehicles):
            vehicle = sim.vehicles[vehicle_id]
            base_comp, base_empty, _ = self._forecast_chain(sim, vehicle, vehicle.queue, now_s)
            base_sum = sum(base_comp)
            for pos in range(len(vehicle.queue) + 1):
                chain = list(vehicle.queue)
                chain.insert(pos, task_id)
                completions, empty_distance, finish = self._forecast_chain(sim, vehicle, chain, now_s)
                new_completion = completions[pos]
                delta_sum = sum(completions) - base_sum
                flow = new_completion - sim.release_s[task_id]
                priority_weight = 1.0 + max(0, task.priority - 50) / 100.0
                score = delta_sum + 0.15 * flow / priority_weight
                score += 1e-5 * (empty_distance - base_empty)
                if (
                    vehicle.current_task is None
                    and not vehicle.queue
                    and pos == 0
                    and not self._idle_task_admissible(sim, vehicle, task)
                ):
                    score += 1_000_000.0
                if self.mode == "balanced":
                    score += 2.0 * len(chain) + 0.8 * self._port_pressure(sim, task)
                    score += 0.02 * max(0.0, finish - now_s)
                candidate = (score, vehicle_id, pos)
                if best is None or candidate < best:
                    best = candidate
        assert best is not None
        return best[1], best[2], best[0]

    def dispatch(
        self, sim: "OnlineTrafficSimulator", ready_ids: Sequence[str], now_s: float
    ) -> List[Dict[str, object]]:
        if self.mode == "document":
            return self._document_auction(sim, ready_ids, now_s)
        if self.mode == "baseline":
            ordered = sorted(ready_ids, key=lambda x: (sim.release_s[x], -self.data.tasks[x].priority, x))
        else:
            ordered = sorted(ready_ids, key=lambda x: (-self.data.tasks[x].priority, sim.release_s[x], x))
        assignments: List[Dict[str, object]] = []
        for task_id in ordered:
            if self.mode == "baseline":
                vehicle_id, pos, score = self._best_baseline(sim, task_id, now_s)
            else:
                vehicle_id, pos, score = self._best_insertion(sim, task_id, now_s)
            sim.assign_task(task_id, vehicle_id, pos, now_s)
            assignments.append({
                "task_id": task_id, "vehicle_id": vehicle_id,
                "queue_position": pos, "bid": score,
            })
        return assignments


class OnlineTrafficSimulator(core.TrafficSimulator):
    QUESTION_NO = 2

    def __init__(
        self,
        data: core.DataBundle,
        graph: core.GraphEngine,
        mode: str,
        capture_trajectory: bool,
    ):
        self.mode = mode
        self.scheduler = OnlineScheduler(data, graph, mode)
        self.reassignment_log: List[Dict[str, object]] = []
        self.reservation_log: List[Dict[str, object]] = []
        self.trigger_log: List[Dict[str, object]] = []
        self.last_reassign_s: Dict[str, float] = {}
        self.reassign_count: Counter[str] = Counter()
        self.pending_fork_trigger = False
        self.pending_resource_trigger = False
        self.pending_deadlock_trigger = False
        self.resource_wait_since: Dict[str, float] = {}
        self.wait_graph_cycle_count = 0
        self.wait_graph_max_edges = 0
        self._last_wait_graph_check_s = -math.inf
        self.base_time = min(t.install_time for t in data.tasks.values())
        self.release_s = {
            task_id: (task.install_time - self.base_time).total_seconds()
            for task_id, task in data.tasks.items()
        }
        carrier_groups: Dict[str, List[core.Task]] = defaultdict(list)
        for task in data.tasks.values():
            carrier_groups[task.carrier_id].append(task)
        self.predecessor: Dict[str, Optional[str]] = {}
        for tasks in carrier_groups.values():
            tasks.sort(key=lambda t: (t.install_time, t.command_id))
            for idx, task in enumerate(tasks):
                self.predecessor[task.command_id] = tasks[idx - 1].command_id if idx else None
        self.release_order = sorted(data.tasks, key=lambda x: (self.release_s[x], x))
        self.release_cursor = 0
        self.released_ids: set[str] = set()
        self.assigned_ids: set[str] = set()
        self.decision_log: List[Dict[str, object]] = []
        self.last_dispatch_s = -math.inf
        super().__init__(data, graph, {v: [] for v in data.vehicles}, mode, capture_trajectory)

    def _route_to_port(self, vehicle: core.Vehicle, port_id: str, kind: str) -> None:
        super()._route_to_port(vehicle, port_id, kind)
        if self.mode == "document" and vehicle.current_task is not None:
            eta = self.time_s + self._remaining_distance(vehicle) / max(1.0, core.CONTROL_VMAX)
            self.reservation_log.append({
                "time_s": round(self.time_s, 3),
                "event": "PORT_INTENT",
                "resource": f"PORT:{port_id}",
                "vehicle_id": vehicle.vehicle_id,
                "task_id": vehicle.current_task,
                "planned_entry_s": round(eta, 3),
                "service_duration_s": core.PICK_TIME if kind == "pickup" else core.DROP_TIME,
            })

    def _cross_boundary(self, vehicle: core.Vehicle, old_link_id: int, next_link_id: int) -> None:
        old_link = self.data.links[old_link_id]
        super()._cross_boundary(vehicle, old_link_id, next_link_id)
        if (
            self.mode == "document"
            and vehicle.current_task is not None
            and bool(vehicle.queue)
            and self.data.nodes[old_link.to_node].role.upper() == "FORK"
        ):
            self.pending_fork_trigger = True

    def _arbitrate_resources(self) -> None:
        old_curve = dict(self.curve_owner)
        old_merge = dict(self.merge_owner)
        super()._arbitrate_resources()
        if self.mode != "document":
            return
        for group, owner in self.curve_owner.items():
            if old_curve.get(group) == owner:
                continue
            vehicle = self.vehicles[owner]
            match = next(
                (x for x in self._upcoming_controls(vehicle) if x[0] == "CURVE" and str(x[1]) == group),
                None,
            )
            self.reservation_log.append({
                "time_s": round(self.time_s, 3), "event": "GRANT",
                "resource": f"CURVE:{group}", "vehicle_id": owner,
                "task_id": vehicle.current_task or "",
                "planned_entry_s": round(self.time_s + (match[2] / max(core.CONTROL_VMAX, vehicle.speed)) if match else self.time_s, 3),
            })
        for node, owner in self.merge_owner.items():
            if old_merge.get(node) == owner:
                continue
            vehicle = self.vehicles[owner]
            match = next(
                (x for x in self._upcoming_controls(vehicle) if x[0] == "MERGE" and int(x[1]) == node),
                None,
            )
            self.reservation_log.append({
                "time_s": round(self.time_s, 3), "event": "GRANT",
                "resource": f"MERGE:{node}", "vehicle_id": owner,
                "task_id": vehicle.current_task or "",
                "planned_entry_s": round(self.time_s + (match[2] / max(core.CONTROL_VMAX, vehicle.speed)) if match else self.time_s, 3),
            })

    @staticmethod
    def _wait_cycle(edges: Dict[str, set[str]]) -> List[str]:
        visiting: set[str] = set()
        visited: set[str] = set()
        stack: List[str] = []

        def visit(node: str) -> Optional[List[str]]:
            visiting.add(node)
            stack.append(node)
            for nxt in edges.get(node, set()):
                if nxt in visiting:
                    i = stack.index(nxt)
                    return stack[i:] + [nxt]
                if nxt not in visited:
                    cycle = visit(nxt)
                    if cycle:
                        return cycle
            stack.pop()
            visiting.remove(node)
            visited.add(node)
            return None

        for node in sorted(edges):
            if node not in visited:
                cycle = visit(node)
                if cycle:
                    return cycle
        return []

    def _update_wait_graph(self) -> None:
        if self.mode != "document" or self.time_s - self._last_wait_graph_check_s < 1.0 - core.EPS:
            return
        edges: Dict[str, set[str]] = defaultdict(set)
        for vehicle in self.vehicles.values():
            if vehicle.state not in self.MOVING_STATES:
                continue
            blocked, _, owner = self._boundary_block(vehicle)
            if blocked and owner and owner != vehicle.vehicle_id:
                edges[vehicle.vehicle_id].add(owner)
            front = self._front_ahead(vehicle)
            if front is not None and front[1] <= 2.0 * core.REF_GAP:
                edges[vehicle.vehicle_id].add(front[0].vehicle_id)
        self.wait_graph_max_edges = max(self.wait_graph_max_edges, sum(len(v) for v in edges.values()))
        cycle = self._wait_cycle(edges)
        if cycle:
            self.wait_graph_cycle_count += 1
            self.pending_deadlock_trigger = True
        self._last_wait_graph_check_s = self.time_s

    def _update_resource_delay_trigger(self) -> None:
        if self.mode != "document":
            return
        waiting_reasons = {"汇流避让", "弯轨等待", "Port等待"}
        active = set()
        for vehicle in self.vehicles.values():
            if vehicle.speed <= core.EPS and vehicle.last_reason in waiting_reasons:
                active.add(vehicle.vehicle_id)
                start = self.resource_wait_since.setdefault(vehicle.vehicle_id, self.time_s)
                if self.time_s - start >= RESOURCE_DELAY_TRIGGER_S - core.EPS:
                    self.pending_resource_trigger = True
            else:
                self.resource_wait_since.pop(vehicle.vehicle_id, None)

    def _reassign_queued_tasks(self, now_s: float) -> List[Dict[str, object]]:
        if self.mode != "document":
            return []
        changes: List[Dict[str, object]] = []
        for old_vehicle in sorted(self.vehicles.values(), key=lambda v: v.vehicle_id):
            for task_id in list(old_vehicle.queue):
                record = self.records[task_id]
                if now_s - record.assigned_s < COMMIT_LOCK_S - core.EPS:
                    continue
                if now_s - self.last_reassign_s.get(task_id, -math.inf) < REASSIGN_COOLDOWN_S - core.EPS:
                    continue
                if self.reassign_count[task_id] >= 2:
                    continue
                old_chain = list(old_vehicle.queue)
                old_pos = old_chain.index(task_id)
                old_completion = self.scheduler._forecast_chain(self, old_vehicle, old_chain, now_s)[0][old_pos]
                old_vehicle.queue.remove(task_id)
                vehicle_id, pos, bid, new_completion = self.scheduler._best_document_insertion(self, task_id, now_s)
                improvement = old_completion - new_completion
                if improvement > REASSIGN_THRESHOLD_S + core.EPS:
                    new_vehicle = self.vehicles[vehicle_id]
                    new_vehicle.queue.insert(pos, task_id)
                    previous_vehicle = record.vehicle_id
                    record.vehicle_id = vehicle_id
                    self.last_reassign_s[task_id] = now_s
                    self.reassign_count[task_id] += 1
                    event = {
                        "time_s": round(now_s, 3), "task_id": task_id,
                        "from_vehicle": previous_vehicle, "to_vehicle": vehicle_id,
                        "new_queue_position": pos,
                        "predicted_improvement_s": round(improvement, 6),
                        "bid": round(bid, 6),
                        "event_type": "跨车重分配" if previous_vehicle != vehicle_id else "同车队列重排",
                    }
                    changes.append(event)
                    self.reassignment_log.append(event)
                else:
                    old_vehicle.queue.insert(old_pos, task_id)
        return changes

    def _initialize(self) -> None:
        for vehicle_id in sorted(self.data.vehicles):
            link_id, offset = self.data.vehicles[vehicle_id]
            vehicle = core.Vehicle(vehicle_id, link_id, offset, [])
            self.vehicles[vehicle_id] = vehicle
            self._activate_next_task(vehicle, 0.0)
        self._release_and_dispatch(0.0, "initial")
        self._handle_arrivals(0.0)
        self._audit_state(0.0)
        self._log_trajectory()

    def _release_tasks(self, now_s: float) -> List[str]:
        released: List[str] = []
        while self.release_cursor < len(self.release_order):
            task_id = self.release_order[self.release_cursor]
            if self.release_s[task_id] > now_s + core.EPS:
                break
            self.released_ids.add(task_id)
            released.append(task_id)
            self.release_cursor += 1
        return released

    def _ready_unassigned(self) -> List[str]:
        ready: List[str] = []
        for task_id in self.released_ids - self.assigned_ids:
            pred = self.predecessor[task_id]
            if pred is None:
                ready.append(task_id)
            elif pred in self.records and self.records[pred].completed_s is not None:
                ready.append(task_id)
        return ready

    def assign_task(self, task_id: str, vehicle_id: str, position: int, now_s: float) -> None:
        if task_id not in self.released_ids:
            raise RuntimeError(f"未来任务被分配: {task_id}")
        pred = self.predecessor[task_id]
        if pred is not None and (pred not in self.records or self.records[pred].completed_s is None):
            raise RuntimeError(f"Carrier 前序未完成: {task_id}")
        vehicle = self.vehicles[vehicle_id]
        self.records[task_id] = core.TaskRecord(self.data.tasks[task_id], vehicle_id, assigned_s=now_s)
        self.assigned_ids.add(task_id)
        vehicle.queue.insert(position, task_id)
        if vehicle.current_task is None:
            self._activate_next_task(vehicle, now_s)

    def _release_and_dispatch(self, now_s: float, trigger: str) -> None:
        new_ids = self._release_tasks(now_s)
        ready = self._ready_unassigned()
        if self.mode == "document":
            heartbeat = now_s - self.last_dispatch_s >= DOCUMENT_HEARTBEAT_S - core.EPS
            high_new = any(self.data.tasks[x].priority >= 90 for x in new_ids)
            event = (
                "release" if new_ids else
                "completion" if trigger == "completion" else
                "fork" if self.pending_fork_trigger else
                "reservation_delay" if self.pending_resource_trigger else
                "deadlock_risk" if self.pending_deadlock_trigger else
                "heartbeat" if heartbeat else ""
            )
            if not event:
                return
            if (
                now_s - self.last_dispatch_s < MIN_REPLAN_INTERVAL_S - core.EPS
                and not high_new
                and event not in {"completion", "deadlock_risk"}
            ):
                return
            max_visible = max((self.release_s[x] for x in self.released_ids), default=-math.inf)
            if max_visible > now_s + core.EPS:
                raise RuntimeError("在线信息泄漏")
            blocked = sorted(self.released_ids - self.assigned_ids - set(ready))
            started = time.perf_counter()
            assignments = self.scheduler.dispatch(self, ready, now_s) if ready else []
            reassignments = self._reassign_queued_tasks(now_s)
            runtime_ms = (time.perf_counter() - started) * 1000.0
            decision = {
                "decision_time_s": round(now_s, 3),
                "trigger": event,
                "new_task_ids": sorted(new_ids),
                "visible_task_ids": sorted(self.released_ids),
                "visible_task_count": len(self.released_ids),
                "blocked_task_ids": blocked,
                "ready_task_ids": sorted(ready),
                "considered_task_ids": sorted(ready),
                "max_visible_release_s": round(max_visible, 3),
                "assignments": assignments,
                "reassignments": reassignments,
                "reservation_count": len(self.curve_owner) + len(self.merge_owner) + len(self.port_owner),
                "solver_runtime_ms": round(runtime_ms, 6),
                "future_leak": False,
            }
            self.decision_log.append(decision)
            self.trigger_log.append({
                "time_s": round(now_s, 3), "trigger": event,
                "new_task_count": len(new_ids), "ready_task_count": len(ready),
            })
            self.last_dispatch_s = now_s
            self.pending_fork_trigger = False
            self.pending_resource_trigger = False
            self.pending_deadlock_trigger = False
            return
        heartbeat = now_s - self.last_dispatch_s >= HEARTBEAT_S - core.EPS
        if not ready or (not new_ids and not heartbeat and trigger != "completion"):
            return
        max_visible = max((self.release_s[x] for x in self.released_ids), default=-math.inf)
        if max_visible > now_s + core.EPS:
            raise RuntimeError("在线信息泄漏")
        assignments = self.scheduler.dispatch(self, ready, now_s)
        self.decision_log.append({
            "decision_time_s": round(now_s, 3),
            "trigger": "release" if new_ids else trigger,
            "new_task_ids": sorted(new_ids),
            "visible_task_count": len(self.released_ids),
            "ready_task_ids": sorted(ready),
            "considered_task_ids": sorted(ready),
            "max_visible_release_s": round(max_visible, 3),
            "assignments": assignments,
            "future_leak": False,
        })
        self.last_dispatch_s = now_s

    def run(self, max_time_s: float = MAX_TIME_S) -> Dict[str, object]:
        started = time.perf_counter()
        previous_completed = self.completed_count
        while self.completed_count < len(self.data.tasks):
            if self.time_s >= max_time_s - core.EPS:
                self._violate("timeout", "", f"{max_time_s}s 未完成全部任务")
                break
            trigger = "completion" if self.completed_count > previous_completed else "heartbeat"
            previous_completed = self.completed_count
            self._release_and_dispatch(self.time_s, trigger)
            self._arbitrate_resources()
            end_s = round(self.time_s + core.DT, 10)
            self._process_services(end_s)

            controls = {}
            for vehicle in self.vehicles.values():
                if vehicle.state not in self.MOVING_STATES:
                    continue
                old_speed = vehicle.speed
                boundary_allowed = not self._boundary_block(vehicle)[0]
                movement, new_speed, reason, related = self._movement_control(vehicle)
                controls[vehicle.vehicle_id] = (
                    movement, new_speed, reason, related, old_speed, boundary_allowed,
                )
            self._audit_continuous_step(controls)
            for vehicle_id, values in controls.items():
                movement, new_speed, reason, related, old_speed, boundary_allowed = values
                vehicle = self.vehicles[vehicle_id]
                self._advance_vehicle(vehicle, movement, new_speed, boundary_allowed)
                vehicle.acceleration = (vehicle.speed - old_speed) / core.DT
                vehicle.last_action = (
                    "停车" if vehicle.speed <= 1e-5
                    else ("减速" if vehicle.acceleration < -1e-5 else "运行")
                )
                vehicle.last_reason = reason
                vehicle.related_vehicle = related
                self._update_pause(vehicle, reason)

            self.time_s = end_s
            self.step_no += 1
            self._handle_arrivals(self.time_s)
            self._update_resource_delay_trigger()
            self._update_wait_graph()
            self._audit_state(self.time_s)
            self._log_trajectory()

        runtime = time.perf_counter() - started
        completed = [r for r in self.records.values() if r.completed_s is not None]
        flows = [r.completed_s - self.release_s[r.task.command_id] for r in completed]
        result = {
            "label": self.mode,
            "completed_tasks": len(completed),
            "makespan_s": max((r.completed_s or 0.0) for r in completed) if completed else math.inf,
            "avg_transfer_time_s": statistics.mean(flows) if flows else math.inf,
            "simulation_runtime_s": runtime,
            "hard_violation_count": len(self.violations),
            "trajectory_rows": len(self.trajectory),
            "decision_count": len(self.decision_log),
            "future_leak_count": sum(bool(x["future_leak"]) for x in self.decision_log),
            "min_same_link_reference_gap_mm": None if math.isinf(self.min_same_link_gap) else self.min_same_link_gap,
            "min_path_reference_gap_mm": None if math.isinf(self.min_path_reference_gap) else self.min_path_reference_gap,
            "min_path_clearance_mm": None if math.isinf(self.min_path_reference_gap) else self.min_path_reference_gap - core.VEHICLE_LENGTH,
            "min_continuous_reference_gap_mm": None if math.isinf(self.min_continuous_reference_gap) else self.min_continuous_reference_gap,
            "min_continuous_clearance_mm": None if math.isinf(self.min_continuous_reference_gap) else self.min_continuous_reference_gap - core.VEHICLE_LENGTH,
            "closest_path_event": self.closest_path_event,
            "closest_continuous_event": self.closest_continuous_event,
            "max_speed_mm_s": self.max_observed_speed,
            "max_acc_mm_s2": self.max_observed_acc,
            "min_acc_mm_s2": self.min_observed_acc,
        }
        result.update({
            "reassignment_count": len(self.reassignment_log),
            "reservation_event_count": len(self.reservation_log),
            "fork_trigger_count": sum(x["trigger"] == "fork" for x in self.trigger_log),
            "reservation_delay_trigger_count": sum(x["trigger"] == "reservation_delay" for x in self.trigger_log),
            "deadlock_risk_trigger_count": sum(x["trigger"] == "deadlock_risk" for x in self.trigger_log),
            "wait_graph_cycle_count": self.wait_graph_cycle_count,
            "wait_graph_max_edges": self.wait_graph_max_edges,
            "decision_runtime_p95_ms": (
                sorted(float(x.get("solver_runtime_ms", 0.0)) for x in self.decision_log)[
                    max(0, math.ceil(0.95 * len(self.decision_log)) - 1)
                ] if self.decision_log else 0.0
            ),
        })
        return result


def absolute_time(base: datetime, relative_s: Optional[float]) -> Optional[datetime]:
    if relative_s is None:
        return None
    return base + timedelta(milliseconds=round(relative_s * 1000.0))


def task_rows(sim: OnlineTrafficSimulator) -> List[List[object]]:
    rows: List[List[object]] = []
    ordered = sorted(sim.data.tasks.values(), key=lambda t: (t.install_time, t.command_id))
    for task in ordered:
        record = sim.records[task.command_id]
        release = sim.release_s[task.command_id]
        if record.completed_s is None:
            raise RuntimeError(f"任务未完成: {task.command_id}")
        rows.append([
            2, task.command_id, task.carrier_id, task.priority, task.install_time,
            task.source, task.destination, record.vehicle_id,
            absolute_time(sim.base_time, record.assigned_s),
            int(round((record.assigned_s - release) * 1000.0)),
            absolute_time(sim.base_time, record.source_arrived_s),
            absolute_time(sim.base_time, record.acquire_start_s),
            absolute_time(sim.base_time, record.acquire_end_s),
            absolute_time(sim.base_time, record.departed_s),
            absolute_time(sim.base_time, record.destination_arrived_s),
            absolute_time(sim.base_time, record.deposit_start_s),
            absolute_time(sim.base_time, record.deposit_end_s),
            absolute_time(sim.base_time, record.completed_s),
            round(record.to_source_distance, 3), round(record.to_destination_distance, 3),
            round(record.completed_s - release, 3), record.move_to_source_path,
            record.move_to_destination_path, record.paused_count, record.pausing_ms,
        ])
    return rows


def metric_row(sim: OnlineTrafficSimulator, runtime_s: float) -> List[object]:
    records = list(sim.records.values())
    flows = [r.completed_s - sim.release_s[r.task.command_id] for r in records]
    return [
        2, 190,
        round(statistics.mean(flows), 6),
        round(statistics.mean(r.assigned_s - sim.release_s[r.task.command_id] for r in records), 6),
        round(statistics.mean(r.source_arrived_s - r.assigned_s for r in records), 6),
        round(statistics.mean(r.destination_arrived_s - r.departed_s for r in records), 6),
        round(statistics.mean(r.pausing_ms / 1000.0 for r in records), 6),
        round(max(r.completed_s for r in records), 6),
        round(statistics.mean(r.to_source_distance for r in records), 6),
        round(statistics.mean(r.to_source_distance + r.to_destination_distance for r in records), 6),
        round(runtime_s, 6),
    ]


def export_excel(
    root: Path,
    out_path: Path,
    tasks: List[List[object]],
    trajectory: List[List[object]],
    metrics: List[object],
) -> None:
    wb = openpyxl.load_workbook(root / "附件9_结果提交模板.xlsx")
    ws_task = wb["任务仿真结果"]
    ws_trace = wb["OHT逐步运行记录表"]
    ws_metric = wb["算法评价指标"]
    if [c.value for c in ws_task[1]] != core.TASK_HEADERS:
        raise ValueError("附件9任务表头不一致")
    q2_rows = [row for row in range(2, ws_task.max_row + 1) if ws_task.cell(row, 1).value == 2]
    if len(q2_rows) != 190:
        raise ValueError(f"附件9问题2预置行 {len(q2_rows)} != 190")
    for row_idx, values in zip(q2_rows, tasks):
        for col_idx, value in enumerate(values, start=1):
            cell = ws_task.cell(row_idx, col_idx, value)
            if isinstance(value, datetime):
                cell.number_format = "yyyy-mm-dd hh:mm:ss.000"
    for row in trajectory:
        if int(row[0]) != 2:
            raise ValueError(f"问题2轨迹问题号错误: {row[0]}")
        ws_trace.append(row)
    for col_idx, value in enumerate(metrics, start=1):
        ws_metric.cell(3, col_idx, value)
    wb.save(out_path)
    check = openpyxl.load_workbook(out_path, data_only=False, read_only=True)
    if check["任务仿真结果"].max_row != 823:
        raise ValueError("任务模板行数被破坏")
    if check["OHT逐步运行记录表"].max_row != len(trajectory) + 1:
        raise ValueError("轨迹写入行数不一致")
    if check["算法评价指标"].cell(3, 2).value != 190:
        raise ValueError("问题2指标未正确写入")
    if any(int(row[0].value) != 2 for row in check["OHT逐步运行记录表"].iter_rows(min_row=2, max_col=1)):
        raise ValueError("问题2 Excel轨迹问题号错误")


def audit_result(
    data: core.DataBundle, sim: OnlineTrafficSimulator, summary: Dict[str, object]
) -> Dict[str, object]:
    errors: List[str] = []
    for task_id in data.tasks:
        if task_id not in sim.records:
            errors.append(f"{task_id}: 未分配")
            continue
        r = sim.records[task_id]
        release = sim.release_s[task_id]
        times = [
            release, r.assigned_s, r.source_arrived_s, r.acquire_start_s,
            r.acquire_end_s, r.departed_s, r.destination_arrived_s,
            r.deposit_start_s, r.deposit_end_s, r.completed_s,
        ]
        if any(x is None for x in times):
            errors.append(f"{task_id}: 时间字段不完整")
            continue
        if any(a > b + core.EPS for a, b in zip(times, times[1:])):
            errors.append(f"{task_id}: 时间非单调")
        if abs((r.acquire_end_s - r.acquire_start_s) - core.PICK_TIME) > 1e-6:
            errors.append(f"{task_id}: 取货时间错误")
        if abs((r.deposit_end_s - r.deposit_start_s) - core.DROP_TIME) > 1e-6:
            errors.append(f"{task_id}: 放货时间错误")
        pred = sim.predecessor[task_id]
        if pred is not None:
            pred_record = sim.records.get(pred)
            if pred_record is None or pred_record.completed_s > r.assigned_s + core.EPS:
                errors.append(f"{task_id}: Carrier 前序约束错误")
    online_errors = []
    for decision in sim.decision_log:
        if decision["max_visible_release_s"] > decision["decision_time_s"] + core.EPS:
            online_errors.append(f"{decision['decision_time_s']}: 读取未来任务")
        if set(decision["considered_task_ids"]) - sim.released_ids:
            online_errors.append(f"{decision['decision_time_s']}: 非法任务ID")
    step_counts = Counter(int(row[1]) for row in sim.trajectory)
    bad_steps = [step for step, count in step_counts.items() if count != 20]
    result = {
        "data_audit": data.audit,
        "simulation_summary": summary,
        "task_validation_errors": errors,
        "online_validation_errors": online_errors,
        "bad_trajectory_steps": bad_steps,
        "traffic_violations": sim.violations,
        "carrier_chain_count": sum(1 for x in sim.predecessor.values() if x is not None),
        "decision_count": len(sim.decision_log),
        "reassignment_audit": {
            "count": len(sim.reassignment_log),
            "events": sim.reassignment_log,
            "errors": [],
        },
        "reservation_audit": {
            "event_count": len(sim.reservation_log),
            "wait_graph_cycle_count": sim.wait_graph_cycle_count,
            "wait_graph_max_edges": sim.wait_graph_max_edges,
        },
    }
    for event in sim.reassignment_log:
        record = sim.records[event["task_id"]]
        if record.source_arrived_s is not None and event["time_s"] >= record.source_arrived_s - core.EPS:
            result["reassignment_audit"]["errors"].append(
                f"{event['task_id']}: 到达Source后仍重分配"
            )
    result["passed"] = (
        summary["completed_tasks"] == 190
        and summary["hard_violation_count"] == 0
        and summary["future_leak_count"] == 0
        and not errors and not online_errors and not bad_steps
        and not result["reassignment_audit"]["errors"]
    )
    return result


def write_report(
    path: Path,
    summaries: Dict[str, Dict[str, object]],
    selected: str,
    metrics: Sequence[object],
) -> None:
    labels = {
        "baseline": "在线FCFS最近完成车辆",
        "rolling": "滚动时域最小增量插入",
        "balanced": "负载均衡滚动拍卖",
        "document": "文档主模型：事件触发滚动拍卖—预约协同",
    }
    lines = [
        "# 第二问：190项动态任务在线调度结果\n",
        "## 模型与简化\n",
        "严格按全过程计划采用事件触发滚动时域拍卖—预约协同。调度器只接收当前已释放且Carrier前序已完成的任务；报价包含预计完成、空载距离、等待、历史拥堵、优先级和年龄。已取货任务冻结，未开始任务仅在超过改善阈值和承诺锁定期后重分配。\n",
        "## 算法比较\n",
        "| 算法 | 平均任务执行时间/s | Makespan/s | 硬违规 |\n",
        "|---|---:|---:|---:|\n",
    ]
    for mode in ["baseline", "rolling", "balanced", "document"]:
        s = summaries[mode]
        lines.append(
            f"| {labels[mode]} | {s['avg_transfer_time_s']:.6f} | {s['makespan_s']:.3f} | {s['hard_violation_count']} |\n"
        )
    lines.extend([
        f"\n最终选择：**{labels[selected]}**。\n",
        "## 正式指标\n",
        f"- AvgTransferTime：{metrics[2]} s\n",
        f"- AvgAssignTime：{metrics[3]} s\n",
        f"- AvgPickupResponseTime：{metrics[4]} s\n",
        f"- AvgTransportTime：{metrics[5]} s\n",
        f"- AvgPausingTime：{metrics[6]} s\n",
        f"- Makespan：{metrics[7]} s\n",
        f"- AvgEmptyDistance：{metrics[8]} mm\n",
        f"- AvgTotalDistance：{metrics[9]} mm\n",
        "\n## 复现\n",
        "```powershell\npython .\\q2_solver.py\npython .\\verify_q2.py\n```\n",
    ])
    path.write_text("".join(lines), encoding="utf-8")


def solve(root: Path) -> Dict[str, object]:
    out = root / "outputs" / "q2"
    out.mkdir(parents=True, exist_ok=True)
    data = load_q2_data(root)
    graph = core.GraphEngine(data)
    summaries: Dict[str, Dict[str, object]] = {}
    for mode in ["baseline", "rolling", "balanced", "document"]:
        print(f"运行问题2 {mode} 无轨迹对比仿真……", flush=True)
        sim = OnlineTrafficSimulator(data, graph, mode, capture_trajectory=False)
        summaries[mode] = sim.run()
        summaries[mode]["hard_gate_passed"] = (
            summaries[mode]["completed_tasks"] == 190
            and summaries[mode]["hard_violation_count"] == 0
            and summaries[mode]["future_leak_count"] == 0
            and summaries[mode]["min_continuous_clearance_mm"] is not None
            and float(summaries[mode]["min_continuous_clearance_mm"]) >= core.CLEAR_GAP - 1e-4
        )
        print(json.dumps(summaries[mode], ensure_ascii=False, indent=2), flush=True)
    feasible = [mode for mode, summary in summaries.items() if summary["hard_gate_passed"] is True]
    if not feasible:
        (out / "问题2_失败诊断.json").write_text(
            json.dumps(summaries, ensure_ascii=False, indent=2), encoding="utf-8"
        )
        raise RuntimeError("问题2候选在线算法均未得到零违规完整解")
    selected = min(
        feasible,
        key=lambda mode: (
            summaries[mode]["avg_transfer_time_s"],
            summaries[mode]["makespan_s"],
            mode,
        ),
    )
    print(f"全日志复跑最终方案：{selected}……", flush=True)
    sim = OnlineTrafficSimulator(data, graph, selected, capture_trajectory=True)
    final_summary = sim.run()
    summaries[selected] = final_summary
    rows = task_rows(sim)
    metrics = metric_row(sim, final_summary["simulation_runtime_s"])
    audit = audit_result(data, sim, final_summary)
    if not audit["passed"]:
        (out / "问题2_失败诊断.json").write_text(
            json.dumps(audit, ensure_ascii=False, indent=2), encoding="utf-8"
        )
        raise RuntimeError("问题2最终方案未通过完整审计")
    stale = out / "问题2_失败诊断.json"
    if stale.exists():
        stale.unlink()

    core.write_csv(out / "问题2_任务结果.csv", core.TASK_HEADERS, rows)
    core.write_csv(out / "问题2_OHT逐步运行记录.csv", core.TRAJECTORY_HEADERS, sim.trajectory)
    core.write_csv(out / "问题2_算法评价指标.csv", core.METRIC_HEADERS, [metrics])
    (out / "问题2_在线决策日志.json").write_text(
        json.dumps(sim.decision_log, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    (out / "问题2_资源预约日志.json").write_text(
        json.dumps(sim.reservation_log, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    (out / "问题2_重分配日志.json").write_text(
        json.dumps(sim.reassignment_log, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    (out / "问题2_约束审计.json").write_text(
        json.dumps(audit, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    plan = {
        "selected_mode": selected,
        "summaries": summaries,
        "configuration": {
            "dt_s": core.DT,
            "heartbeat_s": DOCUMENT_HEARTBEAT_S,
            "control_vmax_mm_s": core.CONTROL_VMAX,
            "clear_gap_mm": core.CLEAR_GAP,
            "reference_gap_mm": core.REF_GAP,
            "minimum_replan_interval_s": MIN_REPLAN_INTERVAL_S,
            "commit_lock_s": COMMIT_LOCK_S,
            "reassignment_improvement_threshold_s": REASSIGN_THRESHOLD_S,
            "reassignment_cooldown_s": REASSIGN_COOLDOWN_S,
            "resource_delay_trigger_s": RESOURCE_DELAY_TRIGGER_S,
            "freeze_after_pickup": True,
            "decision_triggers": ["release", "completion", "fork", "reservation_delay", "deadlock_risk", "heartbeat"],
            "future_task_fields_visible_to_scheduler": False,
            "p17_override": [70, 507, 303.0],
        },
    }
    (out / "问题2_方案.json").write_text(
        json.dumps(plan, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    export_excel(root, out / "问题2_结果.xlsx", rows, sim.trajectory, metrics)
    write_report(out / "问题2_结果说明.md", summaries, selected, metrics)
    return {
        "selected": selected,
        "metrics": dict(zip(core.METRIC_HEADERS, metrics)),
        "audit_passed": True,
        "output_dir": str(out),
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="求解OHT赛题问题2")
    parser.add_argument("--root", type=Path, default=Path(__file__).resolve().parents[1])
    args = parser.parse_args()
    print(json.dumps(solve(args.root.resolve()), ensure_ascii=False, indent=2), flush=True)


if __name__ == "__main__":
    main()
